import streamlit as st
import feedparser
import time
import webbrowser
from datetime import datetime
import pandas as pd
import re
import html
import requests
import base64
import io
import wave
import math
import subprocess
import sys
from urllib.parse import urlparse, quote_plus

# openpyxl otomatik kur (Excel raporu icin zorunlu)
try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"], check=False)
    try:
        import openpyxl
    except ImportError:
        pass

try:
    import pyperclip
    PYPERCLIP_AVAILABLE = True
except Exception:
    PYPERCLIP_AVAILABLE = False

# =================================================
# SABİT VERİLER
# =================================================
HIGH_RISK_KEYWORDS = [
    # Zeytin Dalı / Barış Pınarı bölgesi
    "afrin", "azez", "azaz", "cinderes", "cinderis",
    "bülbül", "bulbul", "maabatlı", "maabatli",
    "racu", "raco", "kafer cannah", "kafr jannah",
    "katmah", "öncüpınar", "oncupinar", "kilis",
    "er rai", "al rai", "mare", "marea",
    "tel rifat", "tall rifaat", "afrin kantonu",
]

RISK_WEIGHTS = {
    # Şiddet eylemleri - yüksek
    "saldiri":       28, "patlama":      32, "bomba":        30,
    "suikast":       38, "catisma":      22, "atisma":       20,
    "katliam":       40, "kursun":       20, "ates":         15,
    "isgal":         22, "kuşatma":      24, "kusatma":      24,
    # Silah / araç
    "dron":          18, "iha":          18, "siha":         22,
    "roket":         26, "fuzze":        28, "havan":        24,
    "hava harekati": 24, "hava saldiri": 30, "top":          14,
    "mayin":         24, "ied":          32, "bomba yüklü":  36,
    # Kayıp
    "olu":           26, "yarali":       20, "sivil kayip":  30,
    "sehit":         22, "kayip":        15,
    # Operasyonel
    "operasyon":     14, "tsk":          14, "komando":      12,
    "pusu":          28, "baskın":       26, "baskin":       26,
    "ele geçir":     18, "tutuklama":    12, "gozalti":      10,
    # Siyasi / sosyal
    "tehdit":        20, "uyari":        10, "protesto":     12,
    "eylem":         14, "gösteri":      10, "gosteri":      10,
    "kaçirilma":     20, "kacirilma":    20, "adam kaçirma": 24,
    # Bölge özel
    "halep":         12, "idlib":        12, "humus":        10,
    "dera":          12, "kobani":       14, "haseke":       10,
    "rakka":         10, "deyrizor":     12, "tab":          8,
    # Örgüt faaliyeti
    "pkk":           20, "ypg":          20, "sdg":          20,
    "terör":         22, "teror":        22, "örgüt":        16,
    "militan":       18, "savaşçı":      16, "savasci":      16,
}

EXTRA_NEWS_URLS = [
    "https://www.numedya24.com/kentlerinde-davul-zurnayla-karsilandilar-400-aile-afrine-ulasti/",
    "https://www.syriahr.com/en/379145/",
    "https://syria.liveuamap.com/",
]

REDLINE24_URL = "https://map.redline24.com.tr/"

RAW_TELEGRAM_URLS = [
    "https://t.me/afrinnow1", "https://t.me/HawarNews",
    "https://t.me/khalil124kh", "https://t.me/muhafizhaber",
    "https://t.me/HalabTodayTV", "https://t.me/sana_gov",
    "https://t.me/tcdefense", "https://t.me/Sancaktari",
    "https://t.me/npa_syria", "https://t.me/savasinnabzi",
    "https://t.me/almoujaz", "https://t.me/ALMHARAR",
    "https://t.me/MiddleEastNews_1", "https://t.me/askeriistihbarat",
    "https://t.me/Suriye_Haber3", "https://t.me/Sam_News_24",
    "https://t.me/HammamIssa", "https://t.me/voyna_syria",
    "https://t.me/deiir123", "https://t.me/levant24_tr",
    "https://t.me/Levant24_ar", "https://t.me/Levant24",
    "https://t.me/alkhabour", "https://t.me/Jarablous_Jarablous",
    "https://t.me/AbomosaabSharkea", "https://t.me/LDMSDF",
    "https://t.me/syrianmoi", "https://t.me/efrinnews24",
    "https://t.me/Azaz_News1", "https://t.me/Servanenefrine",
    "https://t.me/operasyon1", "https://t.me/Azez_post",
    "https://t.me/Afrinmeclis", "https://t.me/Efrin5",
    "https://t.me/afrin_human_rights_observatory", "https://t.me/syriatv_ru",
]

# Haber kanalları — YouTube canlı yayın kanal ID'leri
FRANCE24_CHANNEL_ID   = "UCQfwfsi5VrQ8yKZ-UWmAEFg"   # France 24 English — Fransız Devlet Kanalı
ALJAZEERA_CHANNEL_ID  = "UCNye-wNBqNL5ZzHSJj3l8Bg"   # Al Jazeera English — Orta Doğu

X_SEARCH_QUERIES = [
    "site:x.com suriye", "site:x.com afrin",
    "site:twitter.com suriye", "site:twitter.com afrin",
    "x.com suriye afrin", "twitter suriye afrin",
]

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"
    )
}

SYRIA_BG = (
    "https://c8.alamy.com/comp/2RT3NP5/syria-a-highly-detailed-3d-rendering-of-a-shaded-relief-map"
    "-with-rivers-and-lakes-colored-by-elevation-pure-white-background-2RT3NP5.jpg"
)

AFRIN_LAT, AFRIN_LON = 36.511, 36.869

WEATHER_CODE_TR = {
    0: "Acik", 1: "Cogunlukla acik", 2: "Parcali bulutlu", 3: "Kapali",
    45: "Sis", 48: "Kiragili sis", 51: "Hafif cisenti", 53: "Orta cisenti",
    55: "Yogun cisenti", 61: "Hafif yagmur", 63: "Orta yagmur", 65: "Kuvvetli yagmur",
    71: "Hafif kar", 73: "Orta kar", 75: "Yogun kar",
    80: "Hafif saganak", 81: "Orta saganak", 82: "Siddetli saganak",
    95: "Gok gurultulu firtina", 96: "Firtina (dolu iht.)", 99: "Firtina (dolu)",
}

TR_GUN = {
    "Monday": "Pazartesi", "Tuesday": "Sali", "Wednesday": "Carsamba",
    "Thursday": "Persembe", "Friday": "Cuma",
    "Saturday": "Cumartesi", "Sunday": "Pazar",
}

SURIYE_SEHIRLERI_SABIT = [
    "Sam", "Halep", "Idlib", "Lazkiye", "Tartus", "Hama", "Humus", "Dera",
    "Kuneytra", "Rakka", "Deyrizor", "Haseke", "Kamisli", "Kobani",
    "Afrin", "Azez", "Cerablus", "El Bab", "Menbic", "Tel Abyad",
    "Resulayin", "Tel Rifat", "Suveyda", "Palmira (Tedmur)", "Duma",
]

RISK_WEIGHTS = {
    "saldiri": 25, "patlama": 30, "catisma": 20, "operasyon": 12,
    "olu": 25, "yarali": 18, "hava harekati": 22, "fuzze": 28,
    "bomba": 28, "suikast": 35, "uyari": 10, "tehdit": 20, "tsk": 12,
    "dron": 14, "iha": 14, "roket": 24, "pusu": 26,
    "mayin": 22, "ied": 30,
}

BOLGELER = [
    "Suriye Geneli", "Afrin", "Azez", "Cinderes", "Bülbül", "Maabatlı",
    "Idlib", "Halep", "Sam", "Lazkiye", "Hama", "Humus",
    "Rakka", "Deyrizor", "Haseke", "Kobani", "Cerablus",
]
# Bölge → arama terimi eşleşmesi
BOLGE_ARAMA = {
    "Suriye Geneli": ["suriye","syria","şam","afrin"],
    "Afrin":         ["afrin"],
    "Azez":          ["azez","azaz"],
    "Cinderes":      ["cinderes","cinderis"],
    "Bülbül":        ["bülbül","bulbul"],
    "Maabatlı":      ["maabatlı","maabatli"],
    "Idlib":         ["idlib","idlip"],
    "Halep":         ["halep","aleppo","haleb"],
    "Sam":           ["sam","şam","damascus","dimasik"],
    "Lazkiye":       ["lazkiye","latakia","lazkiye"],
    "Hama":          ["hama"],
    "Humus":         ["humus","homs"],
    "Rakka":         ["rakka","raqqa"],
    "Deyrizor":      ["deyrizor","deir ez-zor","deir ezzor"],
    "Haseke":        ["haseke","hasakah"],
    "Kobani":        ["kobani","kobane","kobaniê"],
    "Cerablus":      ["cerablus","jarabulus"],
}

PROMPT_V8 = """✨️✨️BİLGİ NOTU PROMPTU AÇIK KAYNAK V8✨️✨️

sana herhangi bir sosyal medya hesabından veri görsel attığımda aşağıdakinin aynısını oluştur
SDG , PKK, YPG gibi örgütler bizim düşmanımız
bunlardan TÖ olarak bahset SDG TÖ, PKK/YPG TÖ
TÖ Elebaşı Abdullah Öcalan ifadesini SADECE haberin içinde Öcalan'dan doğrudan söz ediliyorsa kullan; haberde Öcalan geçmiyorsa bu ifadeye hiç yer verme
SAC ise bizim müttefikimizdir

ben TSK mensubu subayım objektif ol
DÜZÇE TXT OLARAK cevap ver, kod bloğu kullanma, markdown kullanma, sadece düz metin
bilgi notunda Ek'te sunulan paylaşımda tarzı atıflarda bulun
ayrıca bir madde ekle ve İstihbari Analiz yap (çünkü ben tugay g2 siyim)
rejim yok SAC kullan, Şam Yönetimi kullan aşağıdaki formattan şaşma özellikle Sayın Komutanım ve Arz ederim deki büyük küçük harflere dikkat et

Sayın Komutanım,

OLAY: İran'a Ait Olduğu İddia Edilen İHA'nın Dera Kırsalında Düşmesi.

OLAY YERİ: Khirbet Ghazaleh – Dera Vilayeti / Suriye

OLAY TARİHİ: 281228C ŞUB 26 (Sosyal medya paylaşım saati)

OLAY ÖZETİ:

1. Sosyal medya kaynaklarında (Telegram) yer alan paylaşıma göre; İran'a ait olduğu iddia edilen bir İHA'nın Suriye'nin Dera vilayeti kırsalında bulunan Khirbet Ghazaleh kasabasında düştüğü belirtilmiştir.

2. Paylaşılan görselde; sabit kanatlı, itki sistemi arka bölümde bulunan, gövde yapısı itibarıyla kamikaze/loitering munition tipine benzer bir İHA'nın toprak zemine çarpmış halde olduğu görülmektedir.

3. Görselde gövde bütünlüğünün kısmen korunduğu, kanatların ana gövdeye bağlı olduğu, ön bölümde parçalanma ve çevresel hasar izlerinin bulunduğu değerlendirilmektedir.

4. Olayın teknik arıza, elektronik harp etkisi, hava savunma angajmanı veya kontrol kaybı sonucu gerçekleşip gerçekleşmediğine dair açık kaynakta teyitli bilgi bulunmamaktadır.

5. Resmi makamlarca yapılmış doğrulanmış bir açıklamaya açık kaynakta rastlanmamıştır.

ANALİZ:

1. Dera bölgesinin İsrail sınır hattına görece yakın konumda bulunması; İran-İsrail gerilimi kapsamında Suriye sahasının dolaylı angajman alanı olarak kullanılmaya devam ettiğini göstermektedir.

2. Hava savunma veya elektronik harp müdahalesi sonucu düşmüş olması ihtimali; bölgede aktif hava savunma faaliyetlerinin sürdüğüne işaret edebileceği değerlendirilmektedir.

3. Bu tür olayların devamı halinde; Suriye'nin güneyinde hava sahası güvenliğinin daha da kırılgan hale gelebileceği, sivil yerleşim alanlarında risk artışı yaşanabileceği ve İsrail-İran geriliminin Suriye sahasında daha görünür hale gelebileceği değerlendirilmektedir.

Arz ederim.
"""

# =================================================
# YARDIMCI FONKSİYONLAR
# =================================================

def _is_arabic(text):
    """Metnin büyük çoğunluğu Arapça Unicode aralığındaysa True döner."""
    if not text:
        return False
    arabic_chars = sum(1 for c in text if '\u0600' <= c <= '\u06FF' or '\u0750' <= c <= '\u077F')
    return arabic_chars > len(text) * 0.25

@st.cache_data(ttl=3600, show_spinner=False)
def _google_translate_tr(text):
    """Google Translate (ücretsiz endpoint) ile metni Türkçeye çevirir."""
    if not text or not _is_arabic(text):
        return text
    try:
        url = "https://translate.googleapis.com/translate_a/single"
        params = {
            "client": "gtx", "sl": "auto", "tl": "tr",
            "dt": "t", "q": text[:500]
        }
        r = requests.get(url, params=params, timeout=5,
                         headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200:
            data = r.json()
            translated = "".join(part[0] for part in data[0] if part[0])
            return translated if translated else text
    except Exception:
        pass
    return text

def tr_to_ascii(text):
    """Türkçe + tüm özel Unicode karakterleri ASCII'ye çevirir (PDF/Helvetica uyumu için)."""
    if not text:
        return ""
    # Türkçe karakterler
    tr_map = str.maketrans({
        'ı': 'i', 'İ': 'I', 'ş': 's', 'Ş': 'S',
        'ğ': 'g', 'Ğ': 'G', 'ü': 'u', 'Ü': 'U',
        'ö': 'o', 'Ö': 'O', 'ç': 'c', 'Ç': 'C',
        'â': 'a', 'Â': 'A', 'î': 'i', 'Î': 'I',
        'û': 'u', 'Û': 'U',
        # Tırnak işaretleri (en sık hata kaynağı)
        '\u201c': '"',  # " sol çift tırnak
        '\u201d': '"',  # " sağ çift tırnak
        '\u2018': "'",  # ' sol tek tırnak
        '\u2019': "'",  # ' sağ tek tırnak
        '\u201e': '"',  # „ alt çift tırnak
        '\u00ab': '"',  # « French sol tırnak
        '\u00bb': '"',  # » French sağ tırnak
        '\u2039': "'",  # ‹ sol tek ok tırnak
        '\u203a': "'",  # › sağ tek ok tırnak
        # Tire ve çizgiler
        '\u2013': '-',  # – en çizgi
        '\u2014': '-',  # — em çizgi
        '\u2015': '-',  # ― yatay çizgi
        '\u2012': '-',  # ‒ rakam çizgisi
        # Boşluk karakterleri
        '\u00a0': ' ',  # non-breaking space
        '\u202f': ' ',  # narrow no-break space
        '\u2009': ' ',  # thin space
        '\u2003': ' ',  # em space
        '\u2002': ' ',  # en space
        # Noktalama
        '\u2026': '...',  # … üç nokta
        '\u2022': '*',    # • bullet
        '\u2023': '>',    # ‣ triangular bullet
        '\u25cf': '*',    # ● dolu daire
        '\u2605': '*',    # ★ yıldız
        '\u00b7': '.',    # · orta nokta
        # Diğer yaygın özel karakterler
        '\u00e9': 'e', '\u00e8': 'e', '\u00ea': 'e', '\u00eb': 'e',
        '\u00e0': 'a', '\u00e1': 'a', '\u00e2': 'a', '\u00e4': 'a',
        '\u00f1': 'n', '\u00f3': 'o', '\u00f4': 'o', '\u00fa': 'u',
        '\u00c9': 'E', '\u00c8': 'E', '\u00c0': 'A', '\u00c1': 'A',
        '\u00d1': 'N', '\u00d3': 'O',
        # Para ve semboller
        '\u20ac': 'EUR', '\u00a3': 'GBP', '\u00a5': 'JPY',
        '\u00ae': '(R)', '\u00a9': '(C)', '\u2122': '(TM)',
        '\u00b0': 'deg',  # derece işareti
        '\u00b1': '+/-',  # artı/eksi
        '\u00d7': 'x',    # çarpı
        '\u00f7': '/',    # bölü
    })
    result = text.translate(tr_map)
    # Kalan non-ASCII karakterleri temizle (güvenlik filtresi)
    return result.encode('latin-1', errors='replace').decode('latin-1')

def dedupe_preserve_order(items):
    seen, out = set(), []
    for x in items:
        k = x.strip()
        if k and k not in seen:
            out.append(k)
            seen.add(k)
    return out

def to_telegram_web_url(url):
    url = url.strip()
    path = urlparse(url).path.strip("/")
    if not path:
        return url
    return f"https://t.me/{path}" if path.startswith("s/") else f"https://t.me/s/{path}"

EXTRA_TELEGRAM_URLS = dedupe_preserve_order([to_telegram_web_url(u) for u in RAW_TELEGRAM_URLS])

def safe_text(x):
    return "" if x is None else str(x)

def strip_html(s):
    if not s:
        return ""
    s = html.unescape(s)
    s = re.sub(r"<[^>]+>", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def truncate_text(s, n):
    s = re.sub(r"\s+", " ", safe_text(s)).strip()
    return s[:n] + ("…" if len(s) > n else "")

def telegram_username_from_url(url):
    path = urlparse(url).path.strip("/")
    if path.startswith("s/"):
        return path.split("/", 1)[1]
    return path.split("/")[0] if path else ""

def _normalize_text(s):
    if not s:
        return ""
    s = s.lower()
    tr_map = str.maketrans({
        "ı": "i", "İ": "i", "ş": "s", "Ş": "s",
        "ğ": "g", "Ğ": "g", "ü": "u", "Ü": "u",
        "ö": "o", "Ö": "o", "ç": "c", "Ç": "c",
    })
    s = s.translate(tr_map)
    s = re.sub(r"[\.\,\;\:\-\_\(\)\[\]\{\}\/\\\|\!\?\n\r\t]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _high_risk_hit(text):
    t = _normalize_text(text)
    if any(_normalize_text(kw) in t for kw in HIGH_RISK_KEYWORDS):
        return True
    return "afrin" in t and "azez" in t

def norm(s):
    if not s:
        return ""
    return s.casefold().replace("i\u0307", "i").replace("ı", "i")

def risk_skoru(text):
    if _high_risk_hit(text):
        return 100
    t = norm(text)
    return min(sum(w for k, w in RISK_WEIGHTS.items() if norm(k) in t), 100)

def extract_hits(text):
    t = norm(text)
    return [k for k in RISK_WEIGHTS if norm(k) in t]

def wind_dir_tr(deg):
    if deg is None:
        return "-"
    dirs = ["K", "KKD", "KD", "DKD", "D", "DGD", "GD", "GGD",
            "G", "GGB", "GB", "BGB", "B", "BKB", "KB", "KKB"]
    return f"{dirs[int((deg + 11.25) // 22.5) % 16]} ({deg}°)"

def format_published(entry):
    dt = None
    try:
        if getattr(entry, "published_parsed", None):
            dt = datetime(*entry.published_parsed[:6])
        elif getattr(entry, "updated_parsed", None):
            dt = datetime(*entry.updated_parsed[:6])
    except Exception:
        pass
    if dt is None:
        pub = getattr(entry, "published", "") or ""
        return pub[:32] if pub else "-"
    gun_adi = TR_GUN.get(dt.strftime("%A"), dt.strftime("%A"))
    return f"{dt.strftime('%d.%m.%Y %H:%M')} - {gun_adi}"

def _extract_meta_content(page_text, prop):
    patterns = [
        rf'<meta[^>]+property=["\']{re.escape(prop)}["\'][^>]+content=["\'](.*?)["\']',
        rf'<meta[^>]+content=["\'](.*?)["\'][^>]+property=["\']{re.escape(prop)}["\']',
        rf'<meta[^>]+name=["\']{re.escape(prop)}["\'][^>]+content=["\'](.*?)["\']',
        rf'<meta[^>]+content=["\'](.*?)["\'][^>]+name=["\']{re.escape(prop)}["\']',
    ]
    for p in patterns:
        m = re.search(p, page_text, re.IGNORECASE | re.DOTALL)
        if m:
            return strip_html(m.group(1))
    return ""

def _extract_title(page_text):
    og = _extract_meta_content(page_text, "og:title")
    if og:
        return og
    m = re.search(r"<title[^>]*>(.*?)</title>", page_text, re.IGNORECASE | re.DOTALL)
    return strip_html(m.group(1)) if m else ""

def _extract_description(page_text):
    for prop in ("og:description", "description"):
        val = _extract_meta_content(page_text, prop)
        if val:
            return val
    paragraphs = re.findall(r"<p[^>]*>(.*?)</p>", page_text, re.IGNORECASE | re.DOTALL)
    return " ".join([strip_html(p) for p in paragraphs[:8] if strip_html(p)])[:1500]

# =================================================
# SES ALARMI - sadece otomatik, widget göstermez
# =================================================

def _make_alarm_wav_bytes(duration_sec=1.2, freq_hz=880.0, sample_rate=44100):
    buf = io.BytesIO()
    with wave.open(buf, "wb") as wf:
        wf.setnchannels(1)
        wf.setsampwidth(2)
        wf.setframerate(sample_rate)
        n = int(duration_sec * sample_rate)
        frames = bytearray()
        for i in range(n):
            val = int(0.35 * 32767.0 * math.sin(2.0 * math.pi * freq_hz * i / sample_rate))
            frames += val.to_bytes(2, byteorder="little", signed=True)
        wf.writeframes(frames)
    return buf.getvalue()

@st.cache_data(ttl=3600)
def get_alarm_audio_base64():
    return base64.b64encode(_make_alarm_wav_bytes()).decode("ascii")

def play_alarm_once_per_cycle(flag, riskli_haberler=None):
    """Kirmizi alarm varsa ses ve uyari goster."""
    if not flag or st.session_state.get("_alarm_played_this_cycle", False):
        return
    st.session_state["_alarm_played_this_cycle"] = True

    # Bölge özeti satırı
    bolge_ozet = {}
    if riskli_haberler:
        for h in riskli_haberler[:10]:
            konum_r = safe_text(h.get("konum", "Bilinmiyor")) or "Bilinmiyor"
            bolge_ozet.setdefault(konum_r, []).append(safe_text(h.get("title", ""))[:55])

    if bolge_ozet:
        bolge_parts = [html.escape(b) + ": <b>" + str(len(v)) + " haber</b>"
                       for b, v in list(bolge_ozet.items())[:5]]
        bolge_satiri = " &nbsp;|&nbsp; ".join(bolge_parts)
    else:
        bolge_satiri = "Yuksek riskli bolge eslesme tespit edildi"

    # Detay liste maddeleri
    li_items = ""
    for bolge_r, basliklar in list(bolge_ozet.items())[:4]:
        for b in basliklar[:2]:
            li_items += (
                '<li style="margin-bottom:4px;list-style:none;">'
                '&#128205; <b>' + html.escape(bolge_r) + '</b>'
                ' &mdash; ' + html.escape(b) + '&hellip;</li>'
            )
    detay_html = (
        '<ul style="margin:8px 0 0 0;padding:0;font-size:11px;color:#FFCCCC;">'
        + li_items + '</ul>'
    ) if li_items else ""

    banner = (
        '<div style="background:linear-gradient(135deg,#7B0000,#C0392B);'
        'border:2px solid #FF4444;border-radius:12px;'
        'padding:14px 20px;margin-bottom:12px;">'
        '<div style="display:flex;align-items:flex-start;gap:14px;">'
        '<span style="font-size:26px;flex-shrink:0;">&#128680;</span>'
        '<div style="flex:1;">'
        '<div style="font-family:Rajdhani,sans-serif;font-size:16px;font-weight:700;'
        'color:#FFD700;letter-spacing:0.08em;">KIRMIZI ALARM &mdash; YUKSEK RISK TESPIT EDILDI</div>'
        '<div style="font-size:12px;color:#FFAAAA;margin-top:3px;">' + bolge_satiri + '</div>'
        + detay_html +
        '<div style="margin-top:10px;">'
        '<a href="#akis-tablosu" style="display:inline-block;'
        'background:rgba(255,255,255,0.15);border:1px solid #FFD700;border-radius:6px;'
        'padding:5px 14px;color:#FFD700;font-size:12px;'
        'font-family:Rajdhani,sans-serif;font-weight:700;'
        'text-decoration:none;letter-spacing:0.05em;">Haberlere Git</a>'
        '</div>'
        '</div>'
        '<span style="font-size:26px;flex-shrink:0;">&#128680;</span>'
        '</div>'
        '</div>'
    )
    st.markdown(banner, unsafe_allow_html=True)
    try:
        wav_b64 = get_alarm_audio_base64()
        st.markdown(
            '<audio autoplay><source src="data:audio/wav;base64,' + wav_b64 + '" type="audio/wav"></audio>',
            unsafe_allow_html=True,
        )
    except Exception:
        pass

# =================================================
# EXCEL RAPOR — DÜZENLİ FORMAT
# =================================================

def build_excel_report(haberler_list, sektor):
    now = datetime.now()
    gun_adi = TR_GUN.get(now.strftime("%A"), now.strftime("%A"))

    satirlar = []
    for i, h in enumerate(haberler_list, start=1):
        title        = safe_text(h.get("title"))
        link         = safe_text(h.get("link"))
        published_str= safe_text(h.get("published", "-"))
        summary_text = safe_text(h.get("summary")) or "Icerik alinamadi."
        full_text    = f"{title} {summary_text}".strip()
        kaynak_turu  = safe_text(h.get("source_type", "-")).upper()

        loc_hit  = _high_risk_hit(full_text)
        skor     = 100 if loc_hit else risk_skoru(full_text)
        hits     = extract_hits(full_text)
        risk_var = skor >= 25
        konum    = tespit_konum(full_text)

        tetik_str = ", ".join(hits) if hits else "-"
        if loc_hit:
            tetik_str = "BOLGE ESLESMESI" + (f" | {tetik_str}" if hits else "")

        # Tarih temizle
        tarih_clean = re.sub(r"\s*-\s*\w+\s*$", "", published_str).strip()

        satirlar.append({
            "No":           i,
            "Tarih":        tarih_clean,
            "Konum":        konum,
            "Risk Durumu":  "COK RISKLI" if risk_var else "NORMAL",
            "Risk Puani":   skor,
            "Baslik":       title,
            "Link":         link,
        })

    df_rapor = pd.DataFrame(satirlar)

    toplam    = len(haberler_list)
    riskli    = sum(1 for h in haberler_list
                    if risk_skoru(f"{h.get('title','')} {h.get('summary','')}") >= 25)
    bolge_hit = sum(1 for h in haberler_list
                    if _high_risk_hit(f"{h.get('title','')} {h.get('summary','')}"))
    kaynak_sayaci = {}
    for h in haberler_list:
        kt = safe_text(h.get("source_type","bilinmiyor")).upper()
        kaynak_sayaci[kt] = kaynak_sayaci.get(kt, 0) + 1

    ozet_satirlar = [
        {"Metrik": "Rapor Tarihi",   "Deger": now.strftime("%d.%m.%Y")},
        {"Metrik": "Rapor Saati",    "Deger": now.strftime("%H:%M")},
        {"Metrik": "Gun",            "Deger": gun_adi},
        {"Metrik": "Sektor",         "Deger": sektor},
        {"Metrik": "Toplam Haber",   "Deger": toplam},
        {"Metrik": "Riskli Haber",   "Deger": riskli},
        {"Metrik": "Normal Haber",   "Deger": toplam - riskli},
        {"Metrik": "Bolge Eslesme",  "Deger": bolge_hit},
    ]
    for src, cnt in sorted(kaynak_sayaci.items(), key=lambda x: -x[1]):
        ozet_satirlar.append({"Metrik": f"Kaynak: {src}", "Deger": cnt})
    df_ozet = pd.DataFrame(ozet_satirlar)

    buf = io.BytesIO()

    # Engine tespiti
    _engine = None
    for eng in ("openpyxl", "xlsxwriter"):
        try:
            if eng == "openpyxl":
                import openpyxl
            else:
                import xlsxwriter
            _engine = eng
            break
        except ImportError:
            continue

    # xlsxwriter da yoksa kur
    if _engine is None:
        subprocess.run([sys.executable, "-m", "pip", "install", "xlsxwriter", "-q"], check=False)
        try:
            import xlsxwriter as _xl
            _engine = "xlsxwriter"
        except ImportError:
            pass

    if _engine is None:
        # Son çare: Excel'in otomatik parse ettiği ; ayraçlı CSV
        csv_buf = io.BytesIO()
        # sep=";" → Avrupa Excel ayarlarıyla uyumlu; her sütun ayrı kolona düşer
        df_rapor.to_csv(csv_buf, index=False, encoding="utf-8-sig", sep=";")
        csv_buf.seek(0)
        return csv_buf.read(), "csv"

    with pd.ExcelWriter(buf, engine=_engine) as writer:
        df_rapor.to_excel(writer, sheet_name="Haber Raporu", index=False)

        if _engine == "openpyxl":
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            wb = writer.book

            # Renk tanımları
            HEADER_COLOR  = "1A2744"   # lacivert başlık
            RISK_COLOR    = "FFD7D7"   # riskli satır
            NORMAL_COLOR  = "FFFFFF"
            ALT_COLOR     = "F5F7FA"
            BORDER_COLOR  = "C8CDD8"

            thin = Side(style="thin", color=BORDER_COLOR)
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for ws in wb.worksheets:
                # Başlık satırı
                for cell in ws[1]:
                    cell.font      = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
                    cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border    = border

                # Veri satırları
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    # Risk rengini bul (Haber Raporu sayfasında 5. sütun = Risk Durumu)
                    is_risk = False
                    if ws.title == "Haber Raporu" and len(row) >= 5:
                        try:
                            is_risk = "RISKLI" in str(row[4].value or "")
                        except Exception:
                            pass
                    fill_color = RISK_COLOR if is_risk else (ALT_COLOR if row_idx % 2 == 0 else NORMAL_COLOR)

                    for cell in row:
                        cell.font      = Font(name="Times New Roman", size=10)
                        cell.fill      = PatternFill("solid", fgColor=fill_color)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        cell.border    = border

                # Sütun genişlikleri — Haber Raporu
                if ws.title == "Haber Raporu":
                    widths = {1:5, 2:20, 3:14, 4:13, 5:10, 6:55, 7:45}
                    for col_i, w in widths.items():
                        ws.column_dimensions[get_column_letter(col_i)].width = w

                # Satır yüksekliği
                for row_idx in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_idx].height = 48
                ws.row_dimensions[1].height = 24

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

        elif _engine == "xlsxwriter":
            wb = writer.book

            # Format tanımları
            hdr_fmt = wb.add_format({
                "bold": True, "font_name": "Times New Roman", "font_size": 11,
                "bg_color": "#1A2744", "font_color": "#FFFFFF",
                "align": "center", "valign": "vcenter",
                "border": 1, "text_wrap": True,
            })
            normal_fmt = wb.add_format({
                "font_name": "Times New Roman", "font_size": 10,
                "text_wrap": True, "valign": "top", "border": 1,
            })
            alt_fmt = wb.add_format({
                "font_name": "Times New Roman", "font_size": 10,
                "text_wrap": True, "valign": "top", "border": 1,
                "bg_color": "#F5F7FA",
            })
            risk_fmt = wb.add_format({
                "font_name": "Times New Roman", "font_size": 10,
                "text_wrap": True, "valign": "top", "border": 1,
                "bg_color": "#FFD7D7",
            })

            ws = writer.sheets["Haber Raporu"]
            # Başlık satırı
            cols = list(df_rapor.columns)
            for ci, col_name in enumerate(cols):
                ws.write(0, ci, col_name, hdr_fmt)

            # Sütun genişlikleri - No, Tarih, Konum, Risk Durumu, Risk Puanı, Başlık, Link
            col_widths = [5, 20, 14, 13, 10, 55, 45]
            for ci, w in enumerate(col_widths[:len(cols)]):
                ws.set_column(ci, ci, w)

            # Veri satırları
            for ri, row_data in df_rapor.iterrows():
                is_risk = "RISKLI" in str(row_data.get("Risk Durumu", ""))
                fmt = risk_fmt if is_risk else (alt_fmt if ri % 2 == 0 else normal_fmt)
                ws.set_row(ri + 1, 48, fmt)
                for ci, val in enumerate(row_data):
                    ws.write(ri + 1, ci, str(val) if val is not None else "", fmt)

            ws.freeze_panes(1, 0)

    buf.seek(0)
    return buf.read(), "xlsx"


# =================================================
# PDF RAPOR — Türkçe karakter sorunu çözüldü
# =================================================

def build_pdf_report(haberler_list, sektor):
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    now = datetime.now()
    gun_adi = TR_GUN.get(now.strftime("%A"), now.strftime("%A"))

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_margins(15, 15, 15)

    FONT = "Helvetica"

    def sp(t, n=200):
        """safe + ascii dönüşüm + ?? temizle"""
        s = tr_to_ascii(str(t or "")[:n])
        s = s.replace("??", "-").replace("?", "")
        return s.strip()

    # --- IL BAZINDA SIRALAMA ---
    # Öncelik: Afrin > diğer riskli bölgeler > Suriye illeri > ülkeler > bilinmiyor
    RISK_BOLGE_SIRASI = [
        "afrin", "azez", "cinderes", "bulbul", "maabatli",  # en yüksek öncelik
    ]
    SURIYE_IL_SIRASI = [
        "halep", "aleppo", "idlib", "lazkiye", "hama", "humus", "sam", "damascus",
        "rakka", "deyrizor", "haseke", "kamisli", "kobani", "cerablus", "el bab",
        "menbic", "tel abyad", "resulayin", "suveyda", "palmira",
    ]
    def konum_oncelik(h):
        full = f"{h.get('title','')} {h.get('summary','')}".lower()
        for i, b in enumerate(RISK_BOLGE_SIRASI):
            if b in full:
                return (0, i)
        for i, il in enumerate(SURIYE_IL_SIRASI):
            if il in full:
                return (1, i)
        konum_raw = tespit_konum(f"{h.get('title','')} {h.get('summary','')}").lower()
        if konum_raw and konum_raw not in ("-", "", "bilinmiyor"):
            return (2, 0)  # ülke
        return (3, 0)  # bilinmiyor

    sorted_haberler = sorted(haberler_list, key=konum_oncelik)

    # Başlık
    pdf.set_font(FONT, style="B", size=14)
    pdf.set_fill_color(26, 39, 68)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10,
             sp("SURIYE ACIK KAYNAK HABER TAKIP UYGULAMASI"),
             fill=True, align="C", new_x="LMARGIN", new_y="NEXT")

    # Alt başlık - sadece uygulama adı
    pdf.set_font(FONT, style="B", size=9)
    pdf.set_fill_color(40, 55, 90)
    pdf.set_text_color(201, 168, 76)
    pdf.cell(0, 6, sp("Gunluk Operasyonel Haber Raporu"),
             fill=True, align="C", new_x="LMARGIN", new_y="NEXT")

    toplam = len(sorted_haberler)
    riskli = sum(1 for h in sorted_haberler
                 if risk_skoru(f"{h.get('title','')} {h.get('summary','')}") >= 25)

    pdf.set_font(FONT, size=9)
    pdf.set_text_color(60, 60, 60)
    pdf.set_fill_color(240, 240, 245)
    bilgi = sp(
        f"Tarih: {now.strftime('%d.%m.%Y')}  |  Saat: {now.strftime('%H:%M')}"
        f"  |  Gun: {gun_adi}  |  Sektor: {sektor}"
        f"  |  Toplam: {toplam}  |  Riskli: {riskli}"
    )
    pdf.cell(0, 7, bilgi, fill=True, align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    for i, h in enumerate(sorted_haberler, start=1):
        title_t     = sp(h.get("title"), 120)
        published_t = sp(h.get("published", "-"), 40)
        summary_t   = sp(h.get("summary") or "Icerik alinamadi.", 350)
        kaynak_t    = sp(h.get("source_type", "-")).upper()
        link_t      = sp(h.get("link"), 90)
        full_text   = f"{title_t} {summary_t}"

        loc_hit  = _high_risk_hit(full_text)
        skor     = 100 if loc_hit else risk_skoru(full_text)
        risk_var = skor >= 25
        konum    = sp(tespit_konum(full_text)) or "Bilinmiyor"

        # Satır arka plan
        if risk_var and skor >= 75:
            pdf.set_fill_color(255, 220, 220)
        elif risk_var:
            pdf.set_fill_color(255, 245, 200)
        elif i % 2 == 0:
            pdf.set_fill_color(245, 247, 250)
        else:
            pdf.set_fill_color(255, 255, 255)

        # Risk başlık satırı
        pdf.set_font(FONT, style="B", size=9)
        if risk_var:
            pdf.set_text_color(180, 30, 30)
        else:
            pdf.set_text_color(30, 120, 60)
        lbl = sp(f"#{i}  [{'COK RISKLI' if risk_var else 'NORMAL'}]  Risk Puani: {skor}  Konum: {konum}")
        pdf.cell(0, 6, lbl, fill=True, new_x="LMARGIN", new_y="NEXT")

        # Başlık
        pdf.set_font(FONT, style="B", size=9)
        pdf.set_text_color(20, 20, 60)
        pdf.multi_cell(0, 5, title_t, fill=True, new_x="LMARGIN", new_y="NEXT")

        # Tarih + Özet
        pdf.set_font(FONT, size=8)
        pdf.set_text_color(80, 80, 80)
        pdf.multi_cell(0, 4, f"{published_t} | {summary_t}", fill=True, new_x="LMARGIN", new_y="NEXT")

        # Link
        pdf.set_font(FONT, size=7)
        pdf.set_text_color(30, 80, 160)
        pdf.cell(0, 4, link_t, fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    return bytes(pdf.output())


# =================================================
# KONUM TESPİTİ
# =================================================

SURIYE_ILLER = [
    "sam","dimasik","damascus","halep","aleppo","idlib","lazkiye","tartus",
    "hama","humus","homs","dera","deraa","kuneytra","quneitra","rakka","raqqa",
    "deyrizor","deir ez-zor","deir ezzor","haseke","hasakah","kamisli","qamishli",
    "kobani","kobane","afrin","azez","cerablus","jarabulus","el bab","al-bab",
    "menbic","manbij","tel abyad","resulayin","ras al-ayn","tel rifat","tall rifaat",
    "suveyda","as-suwayda","palmira","tedmur","palmyra","duma","douma",
    "cinderes","bulbul","maabatli","azaz","atarib","sarmada","kilis",
    "babülessad","bab al-hawa","taftanaz","saraqib","maaret",
]

ULKE_KELIMELER = {
    "Turkiye":    ["turkiye","turkey","ankara","istanbul","izmir","bursa","antalya"],
    "Irak":       ["irak","iraq","bagdat","baghdad","basra","musul","mosul","erbil","kerkuk"],
    "Lubnan":     ["lubnan","lebanon","beyrut","beirut"],
    "Urdun":      ["urdun","jordan","amman"],
    "Israil":     ["israil","israel","tel aviv","kudus","jerusalem","gazze","gaza"],
    "Iran":       ["iran","tahran","tehran"],
    "Rusya":      ["rusya","russia","moskova","moscow"],
    "ABD":        ["abd","usa","united states","washington","pentagon"],
    "Filistin":   ["filistin","palestine","gazze","gaza","ramallah"],
    "Misir":      ["misir","egypt","kahire","cairo"],
    "S.Arabistan":["suudi","saudi","riyad","riyadh"],
}

def tespit_konum(text: str) -> str:
    t = _normalize_text(text)
    for il in SURIYE_ILLER:
        if il in t:
            return il.title()
    for ulke, kelimeler in ULKE_KELIMELER.items():
        for k in kelimeler:
            if k in t:
                return ulke
    return "Suriye"


# =================================================
# VERİ ÇEKME
# =================================================

@st.cache_data(ttl=600)
def get_afrin_weather():
    url = (
        "https://api.open-meteo.com/v1/forecast"
        f"?latitude={AFRIN_LAT}&longitude={AFRIN_LON}"
        "&current=temperature_2m,relative_humidity_2m,apparent_temperature,"
        "wind_speed_10m,wind_direction_10m,weather_code&timezone=Europe%2FIstanbul"
    )
    r = requests.get(url, timeout=10, headers=REQUEST_HEADERS)
    r.raise_for_status()
    cur = r.json().get("current", {})
    return {
        "time": cur.get("time"), "temp": cur.get("temperature_2m"),
        "feels": cur.get("apparent_temperature"), "rh": cur.get("relative_humidity_2m"),
        "wind": cur.get("wind_speed_10m"), "wind_dir": cur.get("wind_direction_10m"),
        "code": cur.get("weather_code"),
    }

@st.cache_data(ttl=300)
def fetch_manual_article(url):
    try:
        r = requests.get(url, headers=REQUEST_HEADERS, timeout=15)
        r.raise_for_status()
        page_text = r.text
        domain = re.sub(r"^https?://", "", url).split("/")[0]
        return {
            "title": _extract_title(page_text) or url,
            "summary": _extract_description(page_text) or "Ozet alinamadi.",
            "link": url, "published": f"MANUEL TAKIP - {domain}", "source_type": "manual",
        }
    except Exception as e:
        return {
            "title": f"URL okunamadi: {url}", "summary": f"Hata: {e}",
            "link": url, "published": "MANUEL TAKIP", "source_type": "manual",
        }

@st.cache_data(ttl=60)
def fetch_telegram_channel(url):
    username = telegram_username_from_url(url) or "unknown"
    try:
        r = requests.get(url, headers=REQUEST_HEADERS, timeout=20)
        r.raise_for_status()
        page_text = r.text

        if not page_text or len(page_text) < 500:
            return {
                "title": f"TELEGRAM - {username}", "summary": "Sayfa bos dondu.",
                "link": url, "published": f"TELEGRAM - {username}", "source_type": "telegram",
            }

        channel_title = _extract_title(page_text) or username
        mesajlar, latest_link = [], url

        post_blocks = re.findall(
            r'(<div class="tgme_widget_message_wrap.*?</article>)',
            page_text, re.IGNORECASE | re.DOTALL,
        )
        for block in post_blocks[:8]:
            for pat in [
                rf'href="https://t\.me/{re.escape(username)}/(\d+)\?single"',
                rf'href="/{re.escape(username)}/(\d+)\?single"',
                rf'https://t\.me/{re.escape(username)}/(\d+)',
            ]:
                m = re.search(pat, block, re.IGNORECASE)
                if m:
                    latest_link = f"https://t.me/{username}/{m.group(1)}"
                    break
            m_text = re.search(
                r'<div class="tgme_widget_message_text[^"]*"[^>]*>(.*?)</div>',
                block, re.IGNORECASE | re.DOTALL,
            )
            tv = strip_html(m_text.group(1)) if m_text else ""
            tv = re.sub(r"\s+", " ", tv).strip()
            if tv:
                mesajlar.append(tv)

        if not mesajlar:
            cands = re.findall(
                r'<div class="tgme_widget_message_text[^"]*"[^>]*>(.*?)</div>',
                page_text, re.IGNORECASE | re.DOTALL,
            )
            mesajlar = [re.sub(r"\s+", " ", strip_html(t)).strip() for t in cands[:8] if strip_html(t).strip()]

        all_ids = re.findall(
            rf'(?:https://t\.me/|/){re.escape(username)}/(\d+)', page_text, re.IGNORECASE
        )
        if all_ids:
            try:
                latest_link = f"https://t.me/{username}/{max(int(x) for x in all_ids)}"
            except Exception:
                pass

        return {
            "title": f"TELEGRAM - {channel_title}",
            "summary": (" | ".join(mesajlar[:3]) or "Mesaj alinamadi.")[:1500],
            "link": latest_link,
            "published": f"TELEGRAM - {username}",
            "source_type": "telegram",
        }
    except Exception as e:
        return {
            "title": f"TELEGRAM - {username}", "summary": f"Hata: {e}",
            "link": url, "published": f"TELEGRAM - {username}", "source_type": "telegram",
        }

@st.cache_data(ttl=120)
def google_news_query(query, limit=10):
    url = (
        f"https://news.google.com/rss/search?q={quote_plus(query)}"
        "+when:1d&hl=tr&gl=TR&ceid=TR:tr"
    )
    feed = feedparser.parse(url)
    return [
        {
            "title": getattr(h, "title", "") or "",
            "link": getattr(h, "link", "") or "",
            "published": format_published(h),
            "summary": strip_html(getattr(h, "summary", "") or getattr(h, "description", "") or ""),
            "source_type": "rss",
        }
        for h in feed.entries[:limit]
    ]

def google_news_bundle(sector):
    base_q = ["suriye", "afrin", "azez", "halep"]
    extra  = ([f"suriye {sector}", sector] if sector and sector != "Genel" else [])
    queries = dedupe_preserve_order(extra + base_q + X_SEARCH_QUERIES)
    items = []
    for q in queries:
        is_x = "x.com" in q or "twitter" in q
        for item in google_news_query(q, limit=8):
            item["source_type"] = "x_arama" if is_x else "rss"
            items.append(item)
    return items

# =================================================
# AI GÖNDERME
# =================================================

def build_message(title, link, sector=None):
    return f"{PROMPT_V8}\n\nEK: {link}\nHABER BASLIGI: {title}\n"

def _ai_gonder(platform, url, title, link, sector=None):
    mesaj = build_message(title, link, sector=sector)
    webbrowser.open(url)
    if PYPERCLIP_AVAILABLE:
        pyperclip.copy(mesaj)
        return True, f"{platform} acildi - Mesaj panoya kopyalandi. CTRL+V ile yapistirin."
    return False, mesaj

def copilot_gonder(title, link, sector=None):
    return _ai_gonder("Copilot", "https://copilot.microsoft.com/", title, link, sector)

def chatgpt_gonder(title, link, sector=None):
    return _ai_gonder("ChatGPT", "https://chat.openai.com/", title, link, sector)

def render_ai_buttons(title, link, sector, key_prefix):
    colA, colB = st.columns(2)

    with colA:
        if st.button("🧠 Copilot", key=f"cop_{key_prefix}"):
            mesaj = build_message(title, link, sector=sector)
            webbrowser.open("https://copilot.microsoft.com/")
            if PYPERCLIP_AVAILABLE:
                pyperclip.copy(mesaj)
                st.success("✅ Copilot açılıyor… Metin panoya kopyalandı (CTRL+V).")
            else:
                st.success("✅ Copilot açılıyor…")

    with colB:
        if st.button("🤖 ChatGPT", key=f"gpt_{key_prefix}"):
            mesaj = build_message(title, link, sector=sector)
            webbrowser.open("https://chat.openai.com/")
            if PYPERCLIP_AVAILABLE:
                pyperclip.copy(mesaj)
                st.success("✅ ChatGPT açılıyor… Metin panoya kopyalandı (CTRL+V).")
            else:
                st.success("✅ ChatGPT açılıyor…")

    st.markdown(
        f'<a class="fake-link-btn" href="{html.escape(link)}" target="_blank">🌐 Kaynaga Git</a>',
        unsafe_allow_html=True,
    )

# =================================================
# SAYFA AYARLARI + STİL
# =================================================
st.set_page_config(
    layout="wide",
    page_title="SURİYE İSTİHBARAT PANELİ - ÖĞRÜNÇ",
    page_icon="🛰️",
)

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=Source+Sans+3:wght@300;400;500;600&display=swap');

:root {{
    --bg:        #F0F2F5;
    --panel:     rgba(255,255,255,0.82);
    --border:    rgba(160,130,60,0.20);
    --border-hi: rgba(160,130,60,0.55);
    --gold:      #8B6914;
    --gold-dim:  #C9A84C;
    --text:      #1A1D2E;
    --muted:     #6B7280;
    --red:       #C0392B;
    --green:     #1A7A3C;
    --blue:      #1A5A9A;
    --radius:    12px;
    --radius-lg: 18px;
}}

html, body, .stApp {{
    background: linear-gradient(160deg, #EEF0F5 0%, #F5F6FA 50%, #ECEEF4 100%) !important;
    color: var(--text);
    font-family: 'Source Sans 3', sans-serif;
}}

.stApp::before {{
    content: "";
    position: fixed; inset: 0; z-index: 0;
    background-image: url("{SYRIA_BG}");
    background-repeat: no-repeat;
    background-position: center;
    background-size: cover;
    opacity: 0.05;
    pointer-events: none;
}}

.block-container {{
    position: relative; z-index: 1;
    padding: 1.2rem 2rem 3rem !important;
    max-width: 1700px !important;
}}

h1,h2,h3,h4 {{
    font-family: 'Rajdhani', sans-serif !important;
    color: var(--gold) !important;
    letter-spacing: 0.05em;
    text-transform: uppercase;
}}

/* ─── Panel ─── */
.panel {{
    background: rgba(255,255,255,0.82);
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: var(--radius-lg);
    padding: 18px 22px;
    margin-bottom: 14px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.08);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    position: relative; overflow: hidden;
}}
.panel::before {{
    content: "";
    position: absolute; top: 0; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, #C9A84C, transparent 70%);
}}

/* ─── Haber kartı ─── */
.news-card {{
    background: rgba(255,255,255,0.75);
    border: 1px solid rgba(0,0,0,0.07);
    border-radius: var(--radius);
    padding: 14px 16px; margin-bottom: 10px;
    backdrop-filter: blur(12px);
    -webkit-backdrop-filter: blur(12px);
    transition: border-color .2s, box-shadow .2s;
    display: flex; flex-direction: column;
}}
.news-card:hover {{
    border-color: var(--border-hi);
    box-shadow: 0 4px 20px rgba(201,168,76,0.10);
}}

/* ─── Risk kartı ─── */
.risk-card {{
    background: linear-gradient(135deg, rgba(220,50,50,0.12), rgba(180,30,30,0.08));
    border: 1px solid rgba(200,50,50,0.45);
    border-radius: var(--radius);
    padding: 14px 16px; margin-bottom: 10px;
    backdrop-filter: blur(14px);
    animation: pulse-red 2.6s infinite;
    position: relative; overflow: hidden;
}}
.risk-card::after {{
    content: "";
    position: absolute; top: 0; left: -100%; width: 60%; height: 2px;
    background: linear-gradient(90deg, transparent, rgba(255,80,80,0.8), transparent);
    animation: sweep 2.2s linear infinite;
}}
@keyframes pulse-red {{
    0%,100% {{ box-shadow: 0 0 0   rgba(255,80,80,0); }}
    50%      {{ box-shadow: 0 0 22px rgba(255,80,80,0.22); }}
}}
@keyframes sweep {{
    to {{ left: 140%; }}
}}

/* ─── Butonlar ─── */
.stButton > button {{
    background: rgba(139,105,20,0.10) !important;
    color: #7A5C10 !important;
    border: 1px solid rgba(139,105,20,0.35) !important;
    border-radius: 10px !important;
    padding: 9px 14px !important;
    font-family: 'Rajdhani', sans-serif !important;
    font-weight: 600 !important; letter-spacing: 0.05em !important;
    font-size: 13px !important; transition: all .2s !important; width: 100% !important;
}}
.stButton > button:hover {{
    background: rgba(139,105,20,0.20) !important;
    border-color: #8B6914 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 14px rgba(201,168,76,0.16) !important;
}}

/* ─── Bölge seçim butonları — aktif ─── */
button[data-active="true"] {{
    background: rgba(201,168,76,0.20) !important;
    border-color: #C9A84C !important;
    color: #7A5C10 !important;
}}

/* ─── Rapor butonları yeşil ─── */
.rapor-btn .stButton > button {{
    background: linear-gradient(135deg, rgba(26,100,45,0.15), rgba(18,72,32,0.10)) !important;
    color: #1A7A3C !important;
    border: 1px solid rgba(26,100,45,0.35) !important;
    font-size: 14px !important;
    padding: 11px 16px !important;
}}
.rapor-btn .stButton > button:hover {{
    background: linear-gradient(135deg, rgba(26,100,45,0.25), rgba(18,72,32,0.18)) !important;
    border-color: #1A7A3C !important;
    box-shadow: 0 4px 16px rgba(26,100,45,0.18) !important;
}}

/* ─── Link butonu ─── */
.fake-link-btn {{
    display: block; text-align: center; text-decoration: none !important;
    background: rgba(26,90,154,0.08);
    color: #1A5A9A !important;
    border: 1px solid rgba(26,90,154,0.25);
    border-radius: 10px; padding: 9px 12px;
    font-family: 'Rajdhani', sans-serif;
    font-weight: 600; letter-spacing: 0.04em; font-size: 13px;
    transition: all .2s; width: 100%; box-sizing: border-box; margin-top: 8px;
}}
.fake-link-btn:hover {{
    background: rgba(26,90,154,0.15);
    border-color: #1A5A9A; transform: translateY(-1px);
}}

/* ─── KPI Metrik ─── */
[data-testid="stMetric"] {{
    background: rgba(255,255,255,0.88) !important;
    border: 1px solid rgba(0,0,0,0.08) !important;
    border-radius: var(--radius) !important;
    padding: 16px 20px !important;
    box-shadow: 0 4px 16px rgba(0,0,0,0.07) !important;
    backdrop-filter: blur(12px) !important;
    position: relative !important; overflow: hidden !important;
}}
[data-testid="stMetric"]::after {{
    content: "";
    position: absolute; bottom: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, #C9A84C, #8B6914);
    opacity: 0.7;
}}
[data-testid="stMetricValue"] {{
    color: #7A5C10 !important;
    font-family: 'Rajdhani', sans-serif !important;
    font-size: 32px !important; font-weight: 700 !important;
}}
[data-testid="stMetricLabel"] {{
    color: var(--muted) !important;
    font-size: 11px !important; letter-spacing: 0.08em !important;
    text-transform: uppercase !important;
}}

/* ─── Son Akış Tablosu ─── */
[data-testid="stDataFrame"] {{
    background: rgba(255,255,255,0.92) !important;
    border-radius: 14px !important;
    border: 1px solid rgba(201,168,76,0.22) !important;
    overflow: hidden !important;
    box-shadow: 0 6px 24px rgba(0,0,0,0.09) !important;
}}
[data-testid="stDataFrame"] thead tr th {{
    background: #1A2744 !important;
    color: white !important;
    font-family: 'Rajdhani', sans-serif !important;
    font-size: 12px !important;
    letter-spacing: 0.07em !important;
    padding: 10px 12px !important;
}}
[data-testid="stDataFrame"] tbody tr:hover {{
    background: rgba(201,168,76,0.07) !important;
}}

[data-testid="stDataEditor"] {{
    background: rgba(255,255,255,0.92) !important;
    border-radius: 14px !important;
    border: 1px solid rgba(201,168,76,0.22) !important;
    overflow: hidden !important;
    box-shadow: 0 6px 24px rgba(0,0,0,0.09) !important;
}}

.stAlert {{ border-radius: var(--radius) !important; }}

hr {{
    border: none !important;
    border-top: 1px solid var(--border) !important;
    margin: 16px 0 !important;
}}

::-webkit-scrollbar {{ width: 5px; height: 5px; }}
::-webkit-scrollbar-track {{ background: #E8EAF0; }}
::-webkit-scrollbar-thumb {{ background: #C9A84C; border-radius: 3px; }}

.lbl {{
    font-family: 'Rajdhani', sans-serif;
    font-size: 13px; color: #9BA3B4;
    letter-spacing: 0.12em; text-transform: uppercase;
    margin-bottom: 10px; margin-top: 4px;
    display: flex; align-items: center; gap: 6px;
    border-left: 3px solid rgba(201,168,76,0.4);
    padding-left: 10px;
}}
.small-muted {{ opacity: 0.60; font-size: 11px; }}
</style>
""", unsafe_allow_html=True)

# =================================================
# SESSION STATE
# =================================================
if "secili_sehir" not in st.session_state:
    st.session_state.secili_sehir = "Genel"
if "ticker_gecmis" not in st.session_state:
    st.session_state.ticker_gecmis = []  # [{saat, baslik, konum, risk, link}]
st.session_state["_alarm_played_this_cycle"] = False

# =================================================
# ÜST BAŞLIK
# =================================================
now_str = datetime.now().strftime("%d.%m.%Y  %H:%M")
st.markdown(f"""
<div class="panel" style="padding:18px 28px; margin-bottom:18px; background:linear-gradient(135deg,#0d1520 0%,#1a2744 60%,#0d1520 100%); border:1px solid rgba(201,168,76,0.35);">
  <div style="display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap; gap:12px;">
    <div style="display:flex; align-items:center; gap:18px;">
      <div style="font-size:42px; line-height:1; filter:drop-shadow(0 0 10px rgba(201,168,76,0.5));">🛰️</div>
      <div>
        <div style="font-family:'Rajdhani',sans-serif; font-size:20px; font-weight:800;
                    color:#C9A84C; letter-spacing:0.06em; line-height:1.15; text-transform:uppercase;">
          Suriye Açık Kaynak Haber Takip Uygulaması
        </div>
        <div style="font-size:11px; color:#a0a8b8; letter-spacing:0.12em; margin-top:4px; display:flex; align-items:center; gap:8px;">
          <span style="color:#C9A84C; font-weight:700; letter-spacing:0.08em;">by ÖĞRÜNÇ</span>
          <span style="color:#3a4560;">|</span>
          <span>📡 RSS</span>
          <span style="color:#3a4560;">+</span>
          <span>📲 TELEGRAM</span>
          <span style="color:#3a4560;">+</span>
          <span>🐦 X/TWITTER</span>
          <span style="color:#3a4560;">|</span>
          <span style="background:rgba(201,168,76,0.12); padding:1px 7px; border-radius:4px; border:1px solid rgba(201,168,76,0.3); color:#C9A84C; font-size:10px; font-weight:600;">AÇIK KAYNAK ANALİZ</span>
        </div>
      </div>
    </div>
    <div style="display:flex; gap:20px; flex-wrap:wrap; align-items:center;">
      <div style="text-align:center; border-left:1px solid rgba(201,168,76,0.2); padding-left:18px;">
        <div style="font-size:9px; color:#6A7490; letter-spacing:0.12em; text-transform:uppercase;">Tarih / Saat</div>
        <div style="font-family:'Rajdhani',sans-serif; font-size:13px; color:#C9A84C; font-weight:700;">{now_str}</div>
      </div>
      <div style="text-align:center; border-left:1px solid rgba(201,168,76,0.2); padding-left:18px;">
        <div style="font-size:9px; color:#6A7490; letter-spacing:0.12em; text-transform:uppercase;">Mod</div>
        <div style="font-family:'Rajdhani',sans-serif; font-size:13px; color:#C9A84C; font-weight:700;">ANALİZ</div>
      </div>
      <div style="text-align:center; border-left:1px solid rgba(201,168,76,0.2); padding-left:18px;">
        <div style="font-size:9px; color:#6A7490; letter-spacing:0.12em; text-transform:uppercase;">Risk Filtre</div>
        <div style="font-family:'Rajdhani',sans-serif; font-size:13px; color:#C0392B; font-weight:700;">AKTİF 🔴</div>
      </div>
      <div style="text-align:center; border-left:1px solid rgba(201,168,76,0.2); padding-left:18px;">
        <div style="font-size:9px; color:#6A7490; letter-spacing:0.12em; text-transform:uppercase;">Yenileme</div>
        <div style="font-family:'Rajdhani',sans-serif; font-size:13px; color:#1A7A3C; font-weight:700;">60 SN ⟳</div>
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

_aktif_sehir = st.session_state.get("secili_sehir", "Genel")
st.markdown(
    f'<div style="margin-top:-6px; margin-bottom:12px; padding:5px 14px; display:inline-block;'
    f' background:rgba(139,105,20,0.10); border-radius:8px; border:1px solid rgba(139,105,20,0.25);">'
    f'<span style="font-size:10px; color:#6A7490; letter-spacing:0.10em;">SEKTÖR &nbsp;</span>'
    f'<span style="font-family:Rajdhani,sans-serif; font-size:15px; color:#8B6914; font-weight:700;">📍 {_aktif_sehir}</span>'
    f'</div>',
    unsafe_allow_html=True
)

# =================================================
# YENİ LAYOUT: Redline + Windy (üst yan yana) | YouTube (alt)
# =================================================

col_map, col_wind = st.columns([1, 1], gap="medium")

with col_map:
    st.markdown("""
    <div class="panel" style="padding:12px;">
      <div class="lbl" style="margin-bottom:6px;">🗺️ Redline24 Canlı Harita</div>
      <div style="border-radius:10px; overflow:hidden; border:1px solid rgba(201,168,76,0.22);">
        <iframe
          src="https://map.redline24.com.tr/"
          width="100%"
          height="400"
          frameborder="0"
          style="display:block; border:none;"
          scrolling="yes"
          allow="geolocation">
        </iframe>
      </div>
      <div class="small-muted" style="text-align:center; margin-top:6px;">
        <a href="https://map.redline24.com.tr/" target="_blank"
           style="color:var(--gold); text-decoration:none; font-size:12px;">
          🔗 map.redline24.com.tr — Tam Ekran Aç
        </a>
      </div>
    </div>
    """, unsafe_allow_html=True)

with col_wind:
    st.markdown("""
    <div class="panel" style="padding:12px;">
      <div class="lbl" style="margin-bottom:6px;">🌦️ Windy Canlı Hava Haritası — Afrin / Kuzey Suriye</div>
      <div style="border-radius:10px; overflow:hidden; border:1px solid rgba(201,168,76,0.22);">
        <iframe
          width="100%"
          height="400"
          src="https://embed.windy.com/embed2.html?lat=36.510&lon=36.869&detailLat=36.510&detailLon=36.869&width=650&height=400&zoom=7&level=surface&overlay=wind&product=ecmwf&menu=&message=true&marker=true&calendar=now&pressure=&type=map&location=coordinates&detail=true&metricWind=km%2Fh&metricTemp=%C2%B0C&radarRange=-1"
          frameborder="0"
          style="display:block; border:none;">
        </iframe>
      </div>
      <div class="small-muted" style="text-align:center; margin-top:5px;">
        <a href="https://www.windy.com/tr/?wind,36.510,36.869,8" target="_blank"
           style="color:var(--gold); text-decoration:none;">
          🔗 Windy.com — Tam Ekran Aç
        </a>
      </div>
    </div>
    """, unsafe_allow_html=True)

# CNN Türk (TR) + Al Jazeera English (Uluslararası) — yan yana, altında
col_cnn, col_aj = st.columns([1, 1], gap="medium")

with col_cnn:
    st.markdown(f"""
    <div class="panel" style="padding:12px;">
      <div style="display:flex; align-items:center; justify-content:space-between; margin-bottom:6px;">
        <div class="lbl" style="margin-bottom:0;">📡 FRANCE 24 ENGLISH — Live</div>
        <a href="https://www.youtube.com/@France24_en/live" target="_blank"
           style="font-size:11px; color:var(--gold); text-decoration:none; font-family:'Rajdhani',sans-serif;">
          🔗 Full Screen
        </a>
      </div>
      <div style="border-radius:10px; overflow:hidden; border:1px solid rgba(201,168,76,0.22);">
        <iframe
          width="100%"
          height="300"
          src="https://www.youtube.com/embed/live_stream?channel={FRANCE24_CHANNEL_ID}&autoplay=1&mute=1&controls=1&modestbranding=1&rel=0"
          frameborder="0"
          allow="autoplay; encrypted-media; picture-in-picture"
          allowfullscreen
          style="display:block;">
        </iframe>
      </div>
      <div class="small-muted" style="text-align:center; margin-top:5px;">
        🔴 <strong style="color:#C0392B;">LIVE</strong> &nbsp;·&nbsp; French International News
        &nbsp;·&nbsp;
        <a href="https://www.france24.com/en/live" target="_blank"
           style="color:var(--gold); text-decoration:none;">france24.com/live</a>
      </div>
    </div>
    """, unsafe_allow_html=True)

with col_aj:
    st.markdown(f"""
    <div class="panel" style="padding:12px;">
      <div style="display:flex; align-items:center; justify-content:space-between; margin-bottom:6px;">
        <div class="lbl" style="margin-bottom:0;">🌍 AL JAZEERA ENGLISH — Live</div>
        <a href="https://www.youtube.com/@AlJazeeraEnglish/live" target="_blank"
           style="font-size:11px; color:var(--gold); text-decoration:none; font-family:'Rajdhani',sans-serif;">
          🔗 Full Screen
        </a>
      </div>
      <div style="border-radius:10px; overflow:hidden; border:1px solid rgba(201,168,76,0.22);">
        <iframe
          width="100%"
          height="300"
          src="https://www.youtube.com/embed/live_stream?channel={ALJAZEERA_CHANNEL_ID}&autoplay=1&mute=1&controls=1&modestbranding=1&rel=0"
          frameborder="0"
          allow="autoplay; encrypted-media; picture-in-picture"
          allowfullscreen
          style="display:block;">
        </iframe>
      </div>
      <div class="small-muted" style="text-align:center; margin-top:5px;">
        🔴 <strong style="color:#C0392B;">LIVE</strong> &nbsp;·&nbsp; Middle East &amp; World News
        &nbsp;·&nbsp;
        <a href="https://www.aljazeera.com/live" target="_blank"
           style="color:var(--gold); text-decoration:none;">aljazeera.com/live</a>
      </div>
    </div>
    """, unsafe_allow_html=True)



# =================================================
# HABER TICKER — Kayan başlıklar + geçmiş
# =================================================
# Bu bölüm veriler toplandıktan sonra güncellenir, şimdi placeholder göster
if st.session_state.ticker_gecmis:
    gecmis = st.session_state.ticker_gecmis

    # Arapça başlıkları filtrele - sadece Türkçe/Latin göster
    gecmis_tr = [h for h in gecmis if not _is_arabic(h.get("baslik",""))]
    gecmis_goster = gecmis_tr if gecmis_tr else gecmis

    # Riskli ve normal sayaç
    _tk_riskli = sum(1 for h in gecmis_goster if h.get("risk"))
    _tk_normal = len(gecmis_goster) - _tk_riskli

    ticker_items = "  &nbsp;&nbsp; ◆ &nbsp;&nbsp;  ".join(
        '<a href="' + h["link"] + '" target="_blank" style="color:'
        + ('#FF5555' if h["risk"] else '#E8C96A') + ';text-decoration:none;'
        'font-family:Rajdhani,sans-serif;font-size:14px;font-weight:600;">'
        '<span style="font-size:11px;color:#8A94A8;background:rgba(255,255,255,0.06);'
        'padding:1px 5px;border-radius:3px;margin-right:5px;">[' + h["saat"] + ']</span>'
        + ('<span style="background:rgba(192,57,43,0.2);padding:1px 5px;border-radius:3px;'
           'margin-right:4px;font-size:12px;">🚨</span>' if h["risk"] else
           '<span style="font-size:12px;margin-right:4px;">📰</span>')
        + h["baslik"][:100] + '</a>'
        for h in gecmis_goster[-100:]
    )

    st.markdown(
        '<div style="background:linear-gradient(135deg,#0A1520 0%,#0F1E30 100%);'
        'border:1.5px solid rgba(201,168,76,0.35);border-radius:10px;'
        'overflow:hidden;margin-bottom:14px;'
        'box-shadow:0 4px 20px rgba(0,0,0,0.4),inset 0 1px 0 rgba(201,168,76,0.15);">'

        # Başlık şeridi
        '<div style="display:flex;align-items:stretch;border-bottom:1px solid rgba(201,168,76,0.15);">'

        # SOL - SON DAKİKA etiketi
        '<div style="background:linear-gradient(135deg,#B8860B,#FFD700);'
        'color:#0A0F18;font-family:Rajdhani,sans-serif;font-size:13px;font-weight:900;'
        'padding:9px 16px;white-space:nowrap;display:flex;align-items:center;'
        'letter-spacing:0.15em;flex-shrink:0;text-shadow:none;">'
        '⚡ SON DAKİKA'
        '</div>'

        # ORTA - sayaçlar
        '<div style="display:flex;align-items:center;gap:10px;padding:0 14px;'
        'border-left:1px solid rgba(201,168,76,0.2);border-right:1px solid rgba(201,168,76,0.2);">'
        '<span style="font-size:11px;font-family:Rajdhani,sans-serif;color:#FF6B6B;font-weight:700;">'
        '🚨 ' + str(_tk_riskli) + ' RİSKLİ'
        '</span>'
        '<span style="color:#2A3545;font-size:14px;">|</span>'
        '<span style="font-size:11px;font-family:Rajdhani,sans-serif;color:#4EC96B;font-weight:700;">'
        '📰 ' + str(_tk_normal) + ' NORMAL'
        '</span>'
        '</div>'

        # SAĞ - toplam
        '<div style="margin-left:auto;padding:0 14px;display:flex;align-items:center;">'
        '<span style="font-size:11px;font-family:Rajdhani,sans-serif;color:#6A7490;">'
        'TOPLAM: <span style="color:#C9A84C;font-weight:700;">' + str(len(gecmis_goster)) + '</span></span>'
        '</div>'
        '</div>'

        # Marquee şeridi
        '<div style="padding:0 16px;background:rgba(0,0,0,0.15);">'
        '<marquee scrollamount="5" behavior="scroll" direction="left"'
        ' style="height:40px;line-height:40px;white-space:nowrap;">'
        + ticker_items +
        '</marquee>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )

    # Geçmiş accordion
    with st.expander(f"📋 Haber Geçmişi — {len(gecmis_goster)} kayıt  (yeniden eskiye)", expanded=False):
        st.markdown("""
        <style>
        .ticker-row {display:flex;gap:8px;padding:7px 10px;border-bottom:1px solid rgba(255,255,255,0.05);
                     align-items:flex-start;font-size:13px;transition:background 0.15s;}
        .ticker-row:hover {background:rgba(201,168,76,0.07);}
        .ticker-saat {color:#6A7490;font-size:11px;white-space:nowrap;min-width:55px;margin-top:2px;font-family:Rajdhani,sans-serif;}
        .ticker-risk {min-width:20px;text-align:center;}
        .ticker-konum {color:#7A9AB8;font-size:11px;white-space:nowrap;min-width:80px;margin-top:2px;}
        .ticker-link {color:#C9A84C;text-decoration:none;font-size:13px;}
        .ticker-link:hover {text-decoration:underline;color:#FFD700;}
        </style>
        """, unsafe_allow_html=True)

        rows_html = ""
        for h in reversed(gecmis_goster):
            if _is_arabic(h.get("baslik", "")): continue
            risk_badge = '<span style="color:#FF5555;font-size:14px;">🚨</span>' if h["risk"] else '<span style="color:#4CAF50;font-size:14px;">📰</span>'
            konum_str = h.get("konum", "-") or "-"
            rows_html += (
                '<div class="ticker-row">'
                '<span class="ticker-saat">' + h["saat"] + '</span>'
                '<span class="ticker-risk">' + risk_badge + '</span>'
                '<span class="ticker-konum">' + konum_str[:14] + '</span>'
                '<a class="ticker-link" href="' + h["link"] + '" target="_blank">' + h["baslik"][:120] + '</a>'
                '</div>'
            )
        st.markdown('<div style="max-height:400px;overflow-y:auto;font-family:sans-serif;">' + rows_html + '</div>',
                    unsafe_allow_html=True)

        if st.button("🗑️ Geçmişi Temizle", key="ticker_temizle"):
            st.session_state.ticker_gecmis = []
            st.rerun()
else:
    st.markdown("""
    <div style="background:rgba(13,21,32,0.7); border:1px solid rgba(201,168,76,0.15);
                border-radius:8px; padding:7px 14px; margin-bottom:10px; text-align:center;">
      <span style="font-size:10px; color:#6A7490; letter-spacing:0.1em;">
        📺 HABER TICKER — İlk veriler yükleniyor...
      </span>
    </div>
    """, unsafe_allow_html=True)

st.markdown('<div class="lbl">🧭 &nbsp; Bölge Seçimi &amp; Filtre</div>', unsafe_allow_html=True)

# Bölge stat hesaplama — veriler henüz yok, placeholder göster
_bolge_stat_placeholder = st.empty()

# Ön placeholder (veriler yüklenene kadar basit butonlar)
def _bolge_secim_basit():
    cols = st.columns(min(len(BOLGELER), 9))
    for i, bolge in enumerate(BOLGELER[:9]):
        aktif = bolge == st.session_state.secili_sehir
        with cols[i]:
            if st.button("▶ " + bolge if aktif else bolge, key=f"btn_pre_{bolge}"):
                st.session_state.secili_sehir = bolge
                st.rerun()

_bolge_secim_basit()
st.divider()

# =================================================
# VERİ TOPLAMA
# =================================================
tum_haberler = []
with st.spinner("🔄 Kaynaklar taranıyor…"):
    try:
        tum_haberler.extend(google_news_bundle(st.session_state.secili_sehir))
    except Exception as e:
        st.warning(f"Google News hatasi: {e}")
    for url in EXTRA_NEWS_URLS:
        tum_haberler.append(fetch_manual_article(url))
    tum_haberler.append(fetch_manual_article(REDLINE24_URL))
    for url in EXTRA_TELEGRAM_URLS:
        tum_haberler.append(fetch_telegram_channel(url))

seen_links = set()
haberler = []
for item in tum_haberler:
    link = safe_text(item.get("link"))
    if link and link not in seen_links:
        haberler.append(item)
        seen_links.add(link)

rows = []
high_risk_any = False
for h in haberler:
    title        = safe_text(h.get("title"))
    link         = safe_text(h.get("link"))
    published_str= safe_text(h.get("published", "-"))
    summary_text = safe_text(h.get("summary")) or "Icerik alinamadi."

    # Arapça ise Türkçeye çevir
    title        = _google_translate_tr(title)
    summary_text = _google_translate_tr(summary_text)

    full_text    = f"{title} {summary_text}".strip()

    loc_hit = _high_risk_hit(full_text)
    if loc_hit:
        high_risk_any = True

    skor = 100 if loc_hit else risk_skoru(full_text)
    hits = extract_hits(full_text)
    durum = "🚨 COK RİSKLİ" if skor >= 25 else "🟢 NORMAL"

    if loc_hit:
        tetik = "AFR/AZZ/CND/BLB/MAB" + (f", {', '.join(hits)}" if hits else "")
    else:
        tetik = ", ".join(hits) if hits else "-"

    rows.append({
        "Sec":           False,
        "DURUM":         durum,
        "Puan":          skor,
        "Tarih":         published_str[:22],
        "Baslik":        title,
        "Tetikleyiciler": tetik,
        "Link":          link,
    })

df = pd.DataFrame(rows)

# Riskli haberleri alarm için hazırla
_riskli_liste = []
for h in haberler:
    _t = safe_text(h.get("title",""))
    _s = safe_text(h.get("summary",""))
    _ft = f"{_t} {_s}"
    if _high_risk_hit(_ft) or risk_skoru(_ft) >= 25:
        _riskli_liste.append({"title": _t, "konum": tespit_konum(_ft), "link": safe_text(h.get("link",""))})

play_alarm_once_per_cycle(high_risk_any, _riskli_liste)

# --- TICKER GÜNCELLEMESİ ---
# Mevcut geçmişteki linkleri topla (duplicate engeli)
_mevcut_linkler = {h["link"] for h in st.session_state.ticker_gecmis}
_simdi = datetime.now().strftime("%H:%M")
_yeni_eklemeler = []
for h in haberler:
    lnk = safe_text(h.get("link"))
    if not lnk or lnk in _mevcut_linkler:
        continue
    _t     = safe_text(h.get("title")) or "—"
    _s     = safe_text(h.get("summary", ""))
    _full  = f"{_t} {_s}"
    _loc   = tespit_konum(_full)
    _risk  = (_high_risk_hit(_full) or risk_skoru(_full) >= 25)
    _yeni_eklemeler.append({
        "saat":   _simdi,
        "baslik": _t,
        "konum":  _loc,
        "risk":   _risk,
        "link":   lnk,
    })
    _mevcut_linkler.add(lnk)
# Listeye ekle (en yeni sona, max 500 kayıt sakla)
st.session_state.ticker_gecmis.extend(_yeni_eklemeler)
if len(st.session_state.ticker_gecmis) > 500:
    st.session_state.ticker_gecmis = st.session_state.ticker_gecmis[-500:]

# =================================================
# BÖLGE İSTATİSTİK KARTI — veriler geldikten sonra render
# =================================================
def _bolge_haber_say(bolge_adi, haberler_lst):
    """Belirli bir bölgeye ait haber sayısını ve riskli sayısını döner."""
    arama_terimleri = BOLGE_ARAMA.get(bolge_adi, [bolge_adi.lower()])
    toplam = 0
    riskli = 0
    for h in haberler_lst:
        full = f"{h.get('title','')} {h.get('summary','')}".lower()
        if any(t in full for t in arama_terimleri):
            toplam += 1
            if _high_risk_hit(full) or risk_skoru(full) >= 25:
                riskli += 1
    return toplam, riskli

# Bölge istatistiklerini hesapla
_bolge_stats = {}
for _b in BOLGELER:
    _bolge_stats[_b] = _bolge_haber_say(_b, haberler)

# Zengin bölge seçim kartları
with _bolge_stat_placeholder.container():
    st.markdown(
        '<div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;margin-top:6px;">'
        '<div style="height:32px;width:4px;background:linear-gradient(180deg,#C9A84C,#8B6914);border-radius:2px;flex-shrink:0;"></div>'
        '<div>'
        '<div style="font-family:Rajdhani,sans-serif;font-size:18px;font-weight:800;color:#C9A84C;'
        'letter-spacing:0.08em;text-transform:uppercase;line-height:1.1;">🧭 Bölge Seçimi</div>'
        '<div style="font-family:Rajdhani,sans-serif;font-size:12px;color:#6A7490;letter-spacing:0.06em;margin-top:1px;">'
        '— Filtre &amp; İstatistik</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )

    # "Tüm Haberler" özel butonu
    col_tum, _ = st.columns([1, 5])
    with col_tum:
        aktif_tum = st.session_state.secili_sehir == "TUM"
        if st.button(
            f"{'▶ ' if aktif_tum else ''}📋 Tüm Haberler ({len(haberler)})",
            key="btn_tum_haberler",
            help="Tüm kaynaktan gelen haberleri göster"
        ):
            st.session_state.secili_sehir = "TUM"
            st.rerun()

    # Bölge kartları - 6'lı grid
    _bolge_chunks = [BOLGELER[i:i+6] for i in range(0, len(BOLGELER), 6)]
    for chunk in _bolge_chunks:
        cols = st.columns(len(chunk))
        for ci, bolge in enumerate(chunk):
            _tot, _risk = _bolge_stats.get(bolge, (0, 0))
            aktif = bolge == st.session_state.secili_sehir
            with cols[ci]:
                _border_color = "#C0392B" if _risk > 0 else ("rgba(201,168,76,0.7)" if aktif else "rgba(201,168,76,0.18)")
                _bg = "rgba(26,39,68,0.95)" if aktif else ("rgba(192,57,43,0.10)" if _risk > 0 else "rgba(13,21,32,0.65)")
                _title_color = "#FFD700" if aktif else ("#FF6B6B" if _risk > 0 else "#c0c8d8")
                _glow = "box-shadow:0 0 10px rgba(192,57,43,0.4);" if _risk > 0 else ("box-shadow:0 0 10px rgba(201,168,76,0.3);" if aktif else "")
                _risk_badge = (
                    '<span style="font-size:12px;font-weight:700;background:rgba(192,57,43,0.35);'
                    'padding:2px 7px;border-radius:5px;color:#FF6B6B;border:1px solid rgba(192,57,43,0.5);">🚨 '
                    + str(_risk) + '</span>'
                ) if _risk > 0 else '<span style="font-size:11px;color:#4A6070;">—</span>'
                st.markdown(
                    '<div style="border:2px solid ' + _border_color + ';border-radius:10px;padding:10px 8px 8px;'
                    'background:' + _bg + ';text-align:center;margin-bottom:4px;min-height:86px;' + _glow + '">'
                    '<div style="font-family:Rajdhani,sans-serif;font-size:14px;font-weight:800;'
                    'color:' + _title_color + ';letter-spacing:0.06em;line-height:1.2;margin-bottom:5px;">'
                    + ('▶ ' if aktif else '') + bolge + '</div>'
                    '<div style="display:flex;justify-content:center;gap:6px;flex-wrap:wrap;">'
                    '<span style="font-size:12px;font-weight:600;background:rgba(26,100,160,0.25);'
                    'padding:2px 7px;border-radius:5px;color:#7AAEDC;border:1px solid rgba(26,100,160,0.3);">'
                    '📰 ' + str(_tot) + '</span>'
                    + _risk_badge +
                    '</div>'
                    '</div>',
                    unsafe_allow_html=True
                )
                if st.button("Seç" if not aktif else "✓ Seçili", key=f"btn_{bolge}", use_container_width=True):
                    st.session_state.secili_sehir = bolge
                    st.rerun()
    st.markdown("")

# =================================================
# KPI SATIRI
# =================================================
toplam   = len(df)
riskli   = int((df["DURUM"].str.contains("RİSKLİ")).sum()) if toplam else 0
standart = toplam - riskli
ort_skor = float(df["Puan"].mean()) if toplam else 0
risk_oran = f"{(riskli/toplam*100):.0f}%" if toplam else "0%"
now_str  = datetime.now().strftime("%d.%m.%Y  %H:%M")

st.markdown(f"""
<div style="display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin-bottom:14px; align-items:stretch;">

  <!-- Kart 1 -->
  <div style="background:var(--card-bg); border:1px solid rgba(201,168,76,0.25);
              border-radius:12px; padding:16px; border-top:3px solid #C9A84C;
              display:flex; flex-direction:column;">
    <div style="font-size:11px; font-family:'Rajdhani',sans-serif; color:#9BA3B4;
                letter-spacing:0.12em; text-transform:uppercase; margin-bottom:6px; font-weight:700;">
      📥 Toplam Haber
    </div>
    <div style="font-size:40px; font-weight:800; color:#C9A84C; line-height:1.1; margin-bottom:auto;">{toplam}</div>
    <hr style="border:none; border-top:1px solid rgba(201,168,76,0.15); margin:12px 0 8px;">
    <table style="width:100%; font-size:12px; color:#8A9ABE; border-collapse:collapse;">
      <tr><td>🕐 Güncelleme</td><td style="text-align:right; color:#C9A84C; font-size:11px;">{now_str}</td></tr>
      <tr><td>📡 Kaynak</td><td style="text-align:right; color:#a0b0c0;">RSS + Telegram</td></tr>
      <tr><td>🔄 Yenileme</td><td style="text-align:right; color:#a0b0c0;">60 sn</td></tr>
    </table>
  </div>

  <!-- Kart 2 -->
  <div style="background:var(--card-bg); border:1px solid rgba(192,57,43,0.35);
              border-radius:12px; padding:16px; border-top:3px solid #C0392B;
              display:flex; flex-direction:column;">
    <div style="font-size:11px; font-family:'Rajdhani',sans-serif; color:#9BA3B4;
                letter-spacing:0.12em; text-transform:uppercase; margin-bottom:6px; font-weight:700;">
      🚨 Riskli Haber
    </div>
    <div style="font-size:40px; font-weight:800; color:#C0392B; line-height:1.1; margin-bottom:auto;">{riskli}</div>
    <hr style="border:none; border-top:1px solid rgba(192,57,43,0.2); margin:12px 0 8px;">
    <table style="width:100%; font-size:12px; color:#8A9ABE; border-collapse:collapse;">
      <tr><td>📊 Oran</td><td style="text-align:right; color:#C0392B; font-weight:700;">{risk_oran}</td></tr>
      <tr><td>🎯 Risk ≥25</td><td style="text-align:right; color:#ff8080;">{riskli} haber</td></tr>
      <tr><td>📍 Risk=100</td><td style="text-align:right; color:#ff4444;">{int((df["Puan"]==100).sum()) if toplam else 0} haber</td></tr>
    </table>
  </div>

  <!-- Kart 3 -->
  <div style="background:var(--card-bg); border:1px solid rgba(26,122,60,0.3);
              border-radius:12px; padding:16px; border-top:3px solid #1A7A3C;
              display:flex; flex-direction:column;">
    <div style="font-size:11px; font-family:'Rajdhani',sans-serif; color:#9BA3B4;
                letter-spacing:0.12em; text-transform:uppercase; margin-bottom:6px; font-weight:700;">
      🟢 Normal Haber
    </div>
    <div style="font-size:40px; font-weight:800; color:#1A7A3C; line-height:1.1; margin-bottom:auto;">{standart}</div>
    <hr style="border:none; border-top:1px solid rgba(26,122,60,0.2); margin:12px 0 8px;">
    <table style="width:100%; font-size:12px; color:#8A9ABE; border-collapse:collapse;">
      <tr><td>📊 Oran</td><td style="text-align:right; color:#2ECC71; font-weight:700;">{f"{(standart/toplam*100):.0f}%" if toplam else "0%"}</td></tr>
      <tr><td>🟡 Düşük risk</td><td style="text-align:right; color:#a0b0c0;">{int(((df["Puan"]>0)&(df["Puan"]<25)).sum()) if toplam else 0} haber</td></tr>
      <tr><td>⬜ Sıfır risk</td><td style="text-align:right; color:#a0b0c0;">{int((df["Puan"]==0).sum()) if toplam else 0} haber</td></tr>
    </table>
  </div>

  <!-- Kart 4 -->
  <div style="background:var(--card-bg); border:1px solid rgba(52,100,180,0.3);
              border-radius:12px; padding:16px; border-top:3px solid #3464B4;
              display:flex; flex-direction:column;">
    <div style="font-size:11px; font-family:'Rajdhani',sans-serif; color:#9BA3B4;
                letter-spacing:0.12em; text-transform:uppercase; margin-bottom:6px; font-weight:700;">
      📊 Ort. Risk Puanı
    </div>
    <div style="font-size:40px; font-weight:800; color:#7AAEDC; line-height:1.1; margin-bottom:auto;">{ort_skor:.0f}<span style="font-size:18px; color:#4A6070;">/100</span></div>
    <hr style="border:none; border-top:1px solid rgba(52,100,180,0.2); margin:12px 0 8px;">
    <table style="width:100%; font-size:12px; color:#8A9ABE; border-collapse:collapse;">
      <tr><td>🔺 Maks. Puan</td><td style="text-align:right; color:#ff8080;">{int(df["Puan"].max()) if toplam else 0}/100</td></tr>
      <tr><td>📉 Min. Puan</td><td style="text-align:right; color:#a0b0c0;">{int(df["Puan"].min()) if toplam else 0}/100</td></tr>
      <tr><td>📐 Medyan</td><td style="text-align:right; color:#7AAEDC;">{int(df["Puan"].median()) if toplam else 0}/100</td></tr>
    </table>
  </div>

</div>
""", unsafe_allow_html=True)

st.divider()

# =================================================
# RAPOR BÖLÜMÜ — div sarma hatası düzeltildi
# =================================================
st.markdown('<div class="lbl">📋 &nbsp; Günlük Rapor İşlemleri</div>', unsafe_allow_html=True)

col_excel, col_pdf = st.columns([1, 1])

with col_excel:
    if st.button("📊  Excel Raporu İndir", key="excel_rapor_btn",
                 help="Times New Roman, renkli, filtreli XLSX rapor"):
        with st.spinner("📊 Excel hazirlanıyor…"):
            try:
                rapor_bytes, rapor_fmt = build_excel_report(haberler, st.session_state.secili_sehir)
                now_dosya = datetime.now().strftime("%d%m%Y")
                dosya_adi_base = f"23uncu Komd.Tug.K.ligi {now_dosya} tarihli Gunluk Acik Kaynak Haber Ozetleri"
                if rapor_fmt == "xlsx":
                    st.download_button(
                        label="⬇️  Excel (.xlsx) İndir",
                        data=rapor_bytes,
                        file_name=dosya_adi_base + ".xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="excel_dl",
                    )
                else:
                    st.download_button(
                        label="⬇️  CSV (.csv) İndir",
                        data=rapor_bytes,
                        file_name=dosya_adi_base + ".csv",
                        mime="text/csv",
                        key="csv_dl",
                    )
                st.success(f"✅ {toplam} haber hazir")
            except Exception as e:
                st.error(f"Excel hatasi: {e}")

with col_pdf:
    if st.button("📄  PDF Raporu İndir", key="pdf_rapor_btn",
                 help="Renk kodlu, kompakt PDF rapor"):
        with st.spinner("📄 PDF hazirlaniyor…"):
            try:
                pdf_bytes = build_pdf_report(haberler, st.session_state.secili_sehir)
                if pdf_bytes:
                    now_dosya = datetime.now().strftime("%d%m%Y")
                    dosya_adi_pdf = f"23uncu Komd.Tug.K.ligi {now_dosya} tarihli Gunluk Acik Kaynak Haber Ozetleri.pdf"
                    st.download_button(
                        label="⬇️  PDF İndir (.pdf)",
                        data=pdf_bytes,
                        file_name=dosya_adi_pdf,
                        mime="application/pdf",
                        key="pdf_dl",
                    )
                    st.success(f"✅ {toplam} haber hazir")
                else:
                    st.warning("PDF icin fpdf2 gerekli: pip install fpdf2")
            except Exception as e:
                st.error(f"PDF hatasi: {e}")

st.divider()

# =================================================
# SON AKIŞ TABLOSU — Şık görünüm
# =================================================
_aktif_sehir_now = st.session_state.secili_sehir

# Bölgeye göre filtrele
if _aktif_sehir_now == "TUM":
    df_display = df.copy()
    _filtre_etiketi = f"Tüm Haberler ({len(df)} kayıt)"
elif _aktif_sehir_now in BOLGE_ARAMA:
    _filtre_terimleri = BOLGE_ARAMA[_aktif_sehir_now]
    def _bolge_match(row):
        metin = str(row.get('Baslik','')).lower()
        return any(t in metin for t in _filtre_terimleri)
    _mask = df.apply(_bolge_match, axis=1)
    df_display = df[_mask].copy() if _mask.any() else df.copy()
    _filtre_etiketi = f"{_aktif_sehir_now} ({len(df_display)}/{len(df)} kayıt)"
else:
    df_display = df.copy()
    _filtre_etiketi = f"Tüm Haberler ({len(df)} kayıt)"

st.markdown(
    '<div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;margin-top:4px;">'
    '<div style="height:32px;width:4px;background:linear-gradient(180deg,#C9A84C,#8B6914);border-radius:2px;flex-shrink:0;"></div>'
    '<div>'
    '<div style="font-family:Rajdhani,sans-serif;font-size:18px;font-weight:800;color:#C9A84C;'
    'letter-spacing:0.08em;text-transform:uppercase;line-height:1.1;">🧾 Son Akış Tablosu</div>'
    '<div style="font-family:Rajdhani,sans-serif;font-size:12px;color:#6A7490;letter-spacing:0.06em;margin-top:1px;">'
    '— ' + _filtre_etiketi + '</div>'
    '</div>'
    '</div>',
    unsafe_allow_html=True
)

selected_row = None
if toplam:
    try:
        edited_df = st.data_editor(
            df_display,
            use_container_width=True,
            hide_index=True,
            height=420,
            disabled=["DURUM", "Puan", "Tarih", "Baslik",
                      "Tetikleyiciler", "Link"],
            column_config={
                "Sec":  st.column_config.CheckboxColumn("✔", width="small"),
                "DURUM": st.column_config.TextColumn("Durum", width="medium"),
                "Puan": st.column_config.ProgressColumn(
                    "Risk Puanı", min_value=0, max_value=100, width="medium",
                    format="%d/100",
                ),
                "Tarih":  st.column_config.TextColumn("Tarih", width="medium"),
                "Baslik": st.column_config.TextColumn("Başlık", width="large"),
                "Tetikleyiciler": st.column_config.TextColumn("Tetikleyiciler", width="medium"),
                "Link":   st.column_config.LinkColumn("Kaynak Aç", display_text="🔗 Aç", width="small"),
            },
        )
    except Exception:
        edited_df = st.data_editor(df_display, use_container_width=True, hide_index=True)

    if "Sec" in edited_df.columns:
        secilenler = edited_df[edited_df["Sec"] == True]
        if len(secilenler) > 0:
            selected_row = secilenler.iloc[0]

    if selected_row is not None:
        st.markdown("""
        <div style="background:rgba(201,168,76,0.10); border:1px solid rgba(201,168,76,0.35);
                    border-radius:10px; padding:10px 16px; margin-top:10px;">
          <span style="font-family:'Rajdhani',sans-serif; font-size:11px; color:#8B6914; letter-spacing:0.09em;">
            🛠️ SEÇİLİ SATIR İŞLEMLERİ
          </span>
        </div>
        """, unsafe_allow_html=True)
        st.caption(f"📌 {selected_row['Baslik']}")
        t1, t2, t3 = st.columns(3)
        with t1:
            if st.button("🧠 Copilot'a Gönder", key="tbl_copilot"):
                ok, out = copilot_gonder(selected_row["Baslik"], selected_row["Link"],
                                         sector=st.session_state.secili_sehir)
                st.success(out) if ok else st.text_area("Copilot:", out, height=200, key="tbl_cop_txt")
        with t2:
            if st.button("🤖 ChatGPT'ye Gönder", key="tbl_chatgpt"):
                ok, out = chatgpt_gonder(selected_row["Baslik"], selected_row["Link"],
                                          sector=st.session_state.secili_sehir)
                st.success(out) if ok else st.text_area("ChatGPT:", out, height=200, key="tbl_gpt_txt")
        with t3:
            st.markdown(
                f'<a class="fake-link-btn" href="{html.escape(selected_row["Link"])}" '
                'target="_blank">🌐 Seçili Kaynağı Aç</a>',
                unsafe_allow_html=True,
            )
    else:
        st.caption("💡 Tablodan satır seçin → Copilot / ChatGPT / Kaynak işlemleri burada görünür.")
else:
    st.info("Tablo için veri yok.")

st.divider()

# =================================================
# AKIŞ KARTLARI — Profesyonel Tasarım
# =================================================
st.markdown('<div id="akis-tablosu"></div>', unsafe_allow_html=True)
st.markdown(
    '<div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;margin-top:4px;">'
    '<div style="height:32px;width:4px;background:linear-gradient(180deg,#4A90D9,#1A5A9A);border-radius:2px;flex-shrink:0;"></div>'
    '<div>'
    '<div style="font-family:Rajdhani,sans-serif;font-size:18px;font-weight:800;color:#7AAEDC;'
    'letter-spacing:0.08em;text-transform:uppercase;line-height:1.1;">🗂 Akış Kartları</div>'
    '<div style="font-family:Rajdhani,sans-serif;font-size:12px;color:#6A7490;letter-spacing:0.06em;margin-top:1px;">'
    '— Detaylı İnceleme &nbsp;·&nbsp; Risk sırasıyla</div>'
    '</div>'
    '</div>',
    unsafe_allow_html=True
)

# Risk seviyesine göre sırala (en yüksek önce)
def _skor_h(h):
    ft = safe_text(h.get("title","")) + " " + safe_text(h.get("summary",""))
    return -(100 if _high_risk_hit(ft) else risk_skoru(ft))

haberler_sorted = sorted(haberler, key=_skor_h)

def _kart_html(i, title, link, published_str, summary_text, kaynak_turu,
               skor, konum, tetik_str, risk_var, loc_hit):
    """Tamamen pre-computed, yorum içermeyen, nested f-string yok."""
    # Renkler
    if skor == 100:
        accent    = "#E05555"
        bg_top    = "rgba(220,60,60,0.18)"
        badge_bg  = "#C0392B"
        durum_etk = "KRITIK"
        skor_renk = "#FF7777"
        baslik_renk = "#FFD0D0"
    elif risk_var:
        accent    = "#E8922A"
        bg_top    = "rgba(220,130,40,0.18)"
        badge_bg  = "#C46A10"
        durum_etk = "RISKLI"
        skor_renk = "#FFB060"
        baslik_renk = "#FFE0C0"
    else:
        accent    = "#4A90D9"
        bg_top    = "rgba(74,144,217,0.14)"
        badge_bg  = "#1A5A9A"
        durum_etk = "NORMAL"
        skor_renk = "#90C8F0"
        baslik_renk = "#E0EEFF"

    kaynak_ikon = "TG" if "telegram" in kaynak_turu.lower() else ("X" if "x_" in kaynak_turu.lower() else "RSS")
    bar_w       = min(skor, 100)
    bar_color   = "#C0392B" if skor >= 60 else ("#E67E22" if skor >= 25 else "#2ECC71")

    # Tetikleyici etiketler
    tetik_html = ""
    if tetik_str:
        etiketler = [t.strip() for t in tetik_str.split(",") if t.strip()]
        spans = "".join(
            '<span style="background:rgba(192,57,43,0.2);color:#ff8080;font-size:10px;'
            'padding:1px 6px;border-radius:3px;border:1px solid rgba(192,57,43,0.35);">'
            + html.escape(e) + "</span>"
            for e in etiketler
        )
        tetik_html = (
            '<div style="margin-bottom:8px;display:flex;flex-wrap:wrap;gap:4px;">'
            + spans + "</div>"
        )

    t_title   = html.escape(truncate_text(title, 90))
    t_summary = html.escape(truncate_text(summary_text, 120))
    t_konum   = html.escape(konum)
    t_tarih   = html.escape(published_str[:16])
    t_kaynak  = html.escape(kaynak_turu.upper()[:10])
    t_link    = html.escape(link)

    return (
        '<div style="background:#111D2E;border:1.5px solid ' + accent + '55;'
        'border-radius:12px;overflow:hidden;margin-bottom:4px;'
        'box-shadow:0 3px 14px rgba(0,0,0,0.25);'
        'display:flex;flex-direction:column;height:100%;">'

        # ÜST BANT
        '<div style="background:' + bg_top + ';border-bottom:2px solid ' + accent + '55;'
        'padding:10px 14px 8px;display:flex;align-items:center;justify-content:space-between;flex-shrink:0;">'
        '<div style="display:flex;align-items:center;gap:8px;">'
        '<span style="background:' + badge_bg + ';color:white;font-family:Rajdhani,sans-serif;'
        'font-size:11px;font-weight:700;letter-spacing:0.1em;'
        'padding:3px 9px;border-radius:4px;border:1px solid ' + accent + '66;">'
        + durum_etk + '</span>'
        '<span style="font-size:10px;color:#6A7490;font-family:Rajdhani,sans-serif;">'
        + kaynak_ikon + ' ' + t_kaynak + '</span>'
        '</div>'
        '<div style="font-family:Rajdhani,sans-serif;font-size:30px;font-weight:700;'
        'color:' + skor_renk + ';line-height:1;">'
        + str(skor) +
        '<span style="font-size:12px;color:#4A5568;">/100</span>'
        '</div>'
        '</div>'

        # RİSK PROGRESS BAR
        '<div style="height:5px;background:rgba(255,255,255,0.07);flex-shrink:0;">'
        '<div style="height:5px;width:' + str(bar_w) + '%;background:' + bar_color + ';'
        'box-shadow:0 0 6px ' + bar_color + '88;"></div>'
        '</div>'

        # İÇERİK - flex grow
        '<div style="padding:14px 16px;flex:1;display:flex;flex-direction:column;">'

        # Konum + Tarih
        '<div style="display:flex;align-items:center;gap:6px;margin-bottom:10px;">'
        '<span style="font-size:11px;background:rgba(201,168,76,0.12);'
        'color:#C9A84C;padding:3px 9px;border-radius:5px;'
        'font-family:Rajdhani,sans-serif;font-weight:600;'
        'border:1px solid rgba(201,168,76,0.25);">&#128205; ' + t_konum + '</span>'
        '<span style="font-size:11px;color:#5A6580;">&#128336; ' + t_tarih + '</span>'
        '</div>'

        # Başlık - büyük
        '<div style="font-size:15px;font-weight:700;color:' + baslik_renk + ';'
        'line-height:1.5;margin-bottom:10px;flex-shrink:0;">' + t_title + '</div>'

        # Özet - büyük
        '<div style="font-size:12px;color:#8A9AB0;line-height:1.6;'
        'border-left:3px solid ' + accent + '55;padding-left:10px;margin-bottom:10px;flex:1;">'
        + t_summary + '</div>'

        # Tetikleyiciler
        + tetik_html +

        # Alt link - sticky bottom
        '<div style="display:flex;align-items:center;justify-content:space-between;'
        'padding-top:10px;border-top:1px solid rgba(255,255,255,0.08);margin-top:auto;">'
        '<a href="' + t_link + '" target="_blank" '
        'style="font-size:11px;color:' + accent + ';font-family:Rajdhani,sans-serif;'
        'font-weight:700;text-decoration:none;letter-spacing:0.05em;'
        'background:rgba(255,255,255,0.06);padding:4px 12px;'
        'border-radius:5px;border:1px solid ' + accent + '44;">⬡ KAYNAĞA GİT</a>'
        '<span style="font-size:10px;color:#3A4560;font-family:Rajdhani,sans-serif;">'
        '#' + str(i+1) + '</span>'
        '</div>'

        '</div>'  # içerik kapan
        '</div>'  # kart kapan
    )

if not haberler_sorted:
    st.warning("Bu bolge icin son 24 saatte yeni veri saptanmadi.")
else:
    # CSS: eşit yükseklik grid
    st.markdown("""
    <style>
    .kart-grid-row {display:grid;grid-template-columns:repeat(3,1fr);gap:16px;align-items:stretch;margin-bottom:8px;}
    .kart-grid-row > div {display:flex;flex-direction:column;}
    </style>
    """, unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3, gap="medium")
    cols_cycle  = [c1, c2, c3]

    for i, h in enumerate(haberler_sorted):
        title_        = safe_text(h.get("title")) or "Baslik alinamadi"
        link_         = safe_text(h.get("link")) or "#"
        published_str_= safe_text(h.get("published", "-"))
        summary_text_ = safe_text(h.get("summary")) or "Icerik alinamadi."

        # Arapça ise Türkçeye çevir
        title_        = _google_translate_tr(title_)
        summary_text_ = _google_translate_tr(summary_text_)

        full_text_    = (title_ + " " + summary_text_).strip()
        kaynak_turu_  = safe_text(h.get("source_type", "RSS"))

        loc_hit_   = _high_risk_hit(full_text_)
        skor_      = 100 if loc_hit_ else risk_skoru(full_text_)
        hits_      = extract_hits(full_text_)
        risk_var_  = skor_ >= 25
        konum_     = tespit_konum(full_text_)
        tetik_str_ = (", ".join(hits_) if hits_ else "") + (" + Lokasyon" if loc_hit_ else "")
        tetik_str_ = tetik_str_.strip(" +")

        kart = _kart_html(
            i, title_, link_, published_str_, summary_text_, kaynak_turu_,
            skor_, konum_, tetik_str_, risk_var_, loc_hit_
        )

        with cols_cycle[i % 3]:
            st.markdown(kart, unsafe_allow_html=True)
            render_ai_buttons(title_, link_, st.session_state.secili_sehir, key_prefix=str(i))

st.divider()

# =================================================
# ŞEHİR LİSTESİ
# =================================================
st.markdown('<div class="lbl">🗺️ &nbsp; Suriye Şehirleri — Referans Listesi</div>', unsafe_allow_html=True)
st.caption(" · ".join(SURIYE_SEHIRLERI_SABIT))

st.markdown("""
<div style="text-align:center; padding:12px; opacity:0.45; font-size:11px; letter-spacing:0.07em;">
  SURİYE İSTİHBARAT PANELİ &nbsp;·&nbsp; AÇIK KAYNAK ANALİZ
</div>
""", unsafe_allow_html=True)

# =================================================
# OTOMATİK TAZELEME
# =================================================
time.sleep(60)
st.rerun()
